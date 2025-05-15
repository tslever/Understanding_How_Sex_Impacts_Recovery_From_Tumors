"""
Melanoma Clinical Data Processing Pipeline
==========================================

This module processes clinical data for melanoma patients from ORIEN datasets.
It implements a clear, step-by-step workflow:

1. Loads essential clinical data files (patient demographics, diagnosis, medications, vital status)
2. Processes and integrates data across these sources
3. Filters for melanoma patients based on histology codes
4. Creates a clean, unified dataset for downstream analysis
5. Generates exploratory plots of demographics and clinical characteristics

The focus is specifically on melanoma patients, and the output is a processed CSV file
with all relevant clinical information including:
- Patient demographics (age, sex, race)
- Cancer diagnosis details (histology, site, stage)
- ICB (immunotherapy) treatment information
- Survival data (OS time, OS status)

Usage:
------
python -m src.data_processing.eda --base-path /path/to/data --output-dir output/melanoma_analysis

Author: Research Team
Created: December 2023
"""

# Standard library imports
import os
import sys
import logging
import argparse
import traceback
import re
import importlib.util
from datetime import datetime
from collections import defaultdict

# Data handling and analysis libraries
import pandas as pd
import numpy as np

# Visualization libraries
import matplotlib.pyplot as plt
import seaborn as sns

# Survival analysis
from lifelines import KaplanMeierFitter
from lifelines.statistics import logrank_test

# Create logger at the module level
logger = logging.getLogger(__name__)

# Set plot style - using a compatible style
plt.style.use('seaborn-v0_8-whitegrid')

# Define ICB drugs for searching in medication records
ICB_DRUGS = {
    'PD-1': ['PEMBROLIZUMAB', 'NIVOLUMAB', 'CEMIPLIMAB', 'SINTILIMAB'],
    'PD-L1': ['ATEZOLIZUMAB', 'AVELUMAB', 'DURVALUMAB'],
    'CTLA-4': ['IPILIMUMAB', 'TREMELIMUMAB'],
    'LAG3': ['RELATLIMAB']
}

# Flatten the ICB drugs list for easier searching
ICB_DRUGS_FLAT = [drug.lower() for sublist in ICB_DRUGS.values() for drug in sublist]

# Define melanoma histology codes from ICD-O-3
# These codes are used to strictly identify melanoma cases
MELANOMA_CODES = [
    '8720', '8721', '8722', '8723', '8728', '8730', '8740',
    '8741', '8742', '8743', '8744', '8745', '8746', '8761',
    '8770', '8771', '8772', '8773', '8774', '8780'
]

# Define ICB Medications (Expand as needed)
ICB_MEDICATIONS = [
    'PEMBROLIZUMAB', 'KEYTRUDA',
    'NIVOLUMAB', 'OPDIVO',
    'ATEZOLIZUMAB', 'TECENTRIQ',
    'DURVALUMAB', 'IMFINZI',
    'AVELUMAB', 'BAVENCIO',
    'IPILIMUMAB', 'YERVOY',
    'CEMIPLIMAB', 'LIBTAYO',
    'TREMELIMUMAB'
]

# ---------------------------
# 1. Data Integration & Cleaning
# ---------------------------

def inspect_dataframe(df, name, verbose=False):
    """Helper function to inspect dataframe structure"""
    if verbose:
        print(f"\n{'-'*50}")
        print(f"Inspecting {name}:")
        print(f"Shape: {df.shape}")
        print("\nColumns:", df.columns.tolist())
        print("\nFirst few rows:")
        print(df.head())
        print("\nMissing values:")
        print(df.isnull().sum().sort_values(ascending=False).head())
        print(f"{'-'*50}\n")
    else:
        print(f"Loaded {name}: {df.shape[0]} rows, {df.shape[1]} columns")

def convert_age(age_str):
    """
    Convert age string to numeric value
    
    Handles special cases:
    - 'Age 90 or older' -> 90.0
    - Non-numeric strings -> np.nan
    - Already numeric values are preserved
    - None/NaN values -> np.nan
    """
    if pd.isna(age_str):
        return np.nan
    if isinstance(age_str, (int, float)):
        return float(age_str)
    if isinstance(age_str, str):
        if 'Age 90 or older' in age_str:
            return 90.0  # Convert censored age to 90
        try:
            return float(age_str)
        except ValueError:
            return np.nan
    return np.nan

def classify_icb(med_name):
    """
    Classify medication name as an immune checkpoint inhibitor (ICB) drug.
    
    Args:
        med_name: Name of the medication
        
    Returns:
        ICB drug name if it's an ICB drug, otherwise None
    """
    if not isinstance(med_name, str):
        return None
        
    # Convert to lowercase for matching
    med_lower = med_name.lower()
    
    # Anti-PD-1 drugs
    if any(drug in med_lower for drug in ['pembrolizumab', 'keytruda']):
        return 'Pembrolizumab'
    if any(drug in med_lower for drug in ['nivolumab', 'opdivo']):
        return 'Nivolumab'
    if any(drug in med_lower for drug in ['cemiplimab', 'libtayo']):
        return 'Cemiplimab'
    if any(drug in med_lower for drug in ['dostarlimab', 'jemperli']):
        return 'Dostarlimab'
    if any(drug in med_lower for drug in ['sintilimab', 'tyvyt']):
        return 'Sintilimab'
    if any(drug in med_lower for drug in ['camrelizumab', 'airuika']):
        return 'Camrelizumab'
    
    # Anti-PD-L1 drugs
    if any(drug in med_lower for drug in ['atezolizumab', 'tecentriq']):
        return 'Atezolizumab'
    if any(drug in med_lower for drug in ['avelumab', 'bavencio']):
        return 'Avelumab'
    if any(drug in med_lower for drug in ['durvalumab', 'imfinzi']):
        return 'Durvalumab'
    
    # Anti-CTLA-4 drugs
    if any(drug in med_lower for drug in ['ipilimumab', 'yervoy']):
        return 'Ipilimumab'
    if any(drug in med_lower for drug in ['tremelimumab', 'imjudo']):
        return 'Tremelimumab'
    
    # LAG-3 inhibitors
    if any(drug in med_lower for drug in ['relatlimab', 'opdualag']):
        return 'Relatlimab'
    
    # Other ICB drugs or combinations
    if any(drug in med_lower for drug in ['pd-1', 'pd1', 'anti-pd-1', 'anti pd-1']):
        return 'PD-1 inhibitor'
    if any(drug in med_lower for drug in ['pd-l1', 'pdl1', 'anti-pd-l1', 'anti pd-l1']):
        return 'PD-L1 inhibitor'
    if any(drug in med_lower for drug in ['ctla-4', 'ctla4', 'anti-ctla-4', 'anti ctla-4']):
        return 'CTLA-4 inhibitor'
    if any(drug in med_lower for drug in ['lag-3', 'lag3', 'anti-lag-3', 'anti lag-3']):
        return 'LAG-3 inhibitor'
    if any(drug in med_lower for drug in ['checkpoint', 'immunotherapy', 'immune checkpoint']):
        return 'Unspecified ICB'
    
    return None

def get_cancer_type(histology_code):
    """
    Determine cancer type from histology code based on strict code list.

    Args:
        histology_code (str): Histology code from ICD-O-3
        
    Returns:
        str: Cancer type (Melanoma or Other)
    """
    # Check if histology code exists and is not NaN
    if pd.isna(histology_code) or histology_code == '':
        return 'Unknown'

    # Extract the base code (first 4 digits) if there's a behavior code
    base_code = str(histology_code).split('/')[0] if '/' in str(histology_code) else str(histology_code)
    base_code = base_code.strip()

    # Match the first 4 digits against melanoma codes
    if any(base_code.startswith(code) for code in MELANOMA_CODES):
        return 'Melanoma'
    else:
        return 'Other'

def get_stage_at_icb(patient_id, base_clinical, icb_start_age, diag_df, outcome_df, stage_input=None):
    """
    Determine the most relevant stage, primary site, and histology for a patient around the time of ICB start,
    OR clean a provided stage value.

    Args:
        patient_id: The PATIENT_ID (used for lookup).
        base_clinical: Row from the main clinical df (used for lookup).
        icb_start_age: Age at first ICB start (used for lookup).
        diag_df: DataFrame with all diagnosis records (used for lookup).
        outcome_df: DataFrame with all outcome records (used for lookup).
        stage_input: If provided, only this stage value will be cleaned.

    Returns:
        If stage_input provided: Tuple (cleaned_stage)
        Otherwise: Tuple (cleaned_stage, relevant_primary_site, relevant_histology)
    """
    stage_to_clean = None
    relevant_primary_site = np.nan  # Initialize site
    relevant_histology = np.nan  # Initialize histology

    # If stage_input is provided, just clean it
    if stage_input is not None:
        stage_to_clean = stage_input
        relevant_primary_site = None 
        relevant_histology = None  # Not relevant when only cleaning stage
    # Otherwise, perform the lookup logic
    elif patient_id is not None:
        # Default values from base_clinical (might be NaN)
        stage_to_clean = base_clinical.get('STAGE', np.nan)
        relevant_primary_site = base_clinical.get('PrimaryDiagnosisSite', np.nan)
        relevant_histology = base_clinical.get('Histology', np.nan)  # Default to base histology if available

        if pd.isna(icb_start_age):
            pass  # Keep defaults if no ICB start age
        else:
            try:
                # Convert icb_start_age to numeric if it's a string
                if isinstance(icb_start_age, str):
                    icb_start_age = float(icb_start_age)
                
                # Filter diagnosis records for the patient
                patient_diag = diag_df[diag_df['PATIENT_ID'] == patient_id].copy()
                
                # Sort by diagnosis age to find the most recent diagnosis before ICB
                if 'AgeAtDiagnosis' in patient_diag.columns:
                    # Convert AgeAtDiagnosis to numeric, coercing errors to NaN
                    patient_diag['AgeAtDiagnosis'] = pd.to_numeric(patient_diag['AgeAtDiagnosis'], errors='coerce')
                    
                    # Get diagnoses before ICB start
                    pre_icb_diag = patient_diag[
                        patient_diag['AgeAtDiagnosis'].notna() &
                        (patient_diag['AgeAtDiagnosis'] <= icb_start_age)
                    ]
                    
                    if not pre_icb_diag.empty:
                        # Get the most recent diagnosis before ICB
                        recent_diag = pre_icb_diag.sort_values('AgeAtDiagnosis', ascending=False).iloc[0]
                        
                        # Get stage information
                        if 'STAGE_PATH' in recent_diag and pd.notna(recent_diag.get('STAGE_PATH')):
                            stage_to_clean = recent_diag['STAGE_PATH']
                        elif 'STAGE_CLIN' in recent_diag and pd.notna(recent_diag.get('STAGE_CLIN')):
                            stage_to_clean = recent_diag['STAGE_CLIN']
                        elif 'TNMClassification_rTNM' in recent_diag and pd.notna(recent_diag.get('TNMClassification_rTNM')):
                            stage_to_clean = recent_diag['TNMClassification_rTNM']
                        
                        # Update site and histology
                        if 'PrimaryDiagnosisSite' in recent_diag and pd.notna(recent_diag.get('PrimaryDiagnosisSite')):
                            relevant_primary_site = recent_diag['PrimaryDiagnosisSite']
                        if 'Histology' in recent_diag and pd.notna(recent_diag.get('Histology')):
                            relevant_histology = recent_diag['Histology']
                    else:
                        # No diagnosis before ICB, try to use information from any diagnosis
                        if not patient_diag.empty:
                            # Use the earliest diagnosis as fallback
                            earliest_diag = patient_diag.sort_values('AgeAtDiagnosis').iloc[0]
                            
                            # Get stage information
                            if 'STAGE_PATH' in earliest_diag and pd.notna(earliest_diag.get('STAGE_PATH')):
                                stage_to_clean = earliest_diag['STAGE_PATH']
                            elif 'STAGE_CLIN' in earliest_diag and pd.notna(earliest_diag.get('STAGE_CLIN')):
                                stage_to_clean = earliest_diag['STAGE_CLIN']
                                
                            # Update site and histology
                            if 'PrimaryDiagnosisSite' in earliest_diag and pd.notna(earliest_diag.get('PrimaryDiagnosisSite')):
                                relevant_primary_site = earliest_diag['PrimaryDiagnosisSite']
                            if 'Histology' in earliest_diag and pd.notna(earliest_diag.get('Histology')):
                                relevant_histology = earliest_diag['Histology']
            except Exception as e:
                logger.error(f"Error determining stage at ICB for patient {patient_id}: {e}")
                stage_to_clean = 'Error'
    else: 
        stage_to_clean = 'Unknown'
        relevant_primary_site = 'Unknown'
        relevant_histology = 'Unknown'

    # Cleaning logic (now common path)
    cleaned_stage = clean_stage_value(stage_to_clean)

    # If called for cleaning only, return just the stage
    if stage_input is not None:
        return cleaned_stage
    # Otherwise return stage, site, and histology
    else:
        return cleaned_stage, relevant_primary_site, relevant_histology

def clean_stage_value(stage):
    """
    Clean stage values to standard format.

    Args:
        stage (str): Raw stage value

    Returns:
        str: Cleaned stage value
    """
    if pd.isna(stage) or stage == '' or 'unknown' in str(stage).lower():
        return 'Unknown'

    # Remove prefix characters and whitespace
    stage_clean = re.sub(r'^[cp]', '', str(stage)).strip()

    # Map to standard stage values
    stage_map = {
        '0': 'Stage 0',
        'I': 'Stage I', 'IA': 'Stage IA', 'IB': 'Stage IB', 'IS': 'Stage IS',
        'II': 'Stage II', 'IIA': 'Stage IIA', 'IIB': 'Stage IIB', 'IIC': 'Stage IIC',
        'III': 'Stage III', 'IIIA': 'Stage IIIA', 'IIIB': 'Stage IIIB', 'IIIC': 'Stage IIIC', 'IIID': 'Stage IIID',
        'IV': 'Stage IV', 'IVA': 'Stage IVA', 'IVB': 'Stage IVB', 'IVC': 'Stage IVC'
    }

    return stage_map.get(stage_clean, stage_clean if stage_clean else 'Unknown')

def load_clinical_data(base_path):
    """
    Load only the essential clinical data files needed for processing.
    
    Args:
        base_path (str): Base path to the project directory
    
    Returns:
        dict: Dictionary of dataframes with clinical data
    """
    logger.info("==================== STARTING DATA LOADING ====================")
    logger.info("Loading clinical data files from normalized directory")

    # Hard-code the expected path for normalized files
    normalized_path = os.path.join(base_path, 'Clinical_Data', '24PRJ217UVA_NormalizedFiles')
    logger.info(f"Looking for files in: {normalized_path}")

    # Define only the essential files we need for melanoma analysis
    essential_files = {
        'patients': 'PatientMaster_V4.csv',
        'diagnosis': 'Diagnosis_V4.csv',
        'medications': 'Medications_V4.csv',
        'vital': 'VitalStatus_V4.csv',
        'outcomes': 'Outcomes_V4.csv',  # Add outcomes data for progression/recurrence information
        'tumor_sequencing': 'TumorSequencing_V4.csv'  # Add tumor sequencing data
    }

    # Initialize dictionary to store dataframes
    dfs = {}

    # Check if normalized path exists
    if not os.path.exists(normalized_path):
        logger.error(f"ERROR: Normalized files directory not found: {normalized_path}")
        logger.error("Please check that the base path is correct")
        return dfs
    
    # Get a listing of all files in the directory
    try:
        all_files = os.listdir(normalized_path)
        logger.info(f"Found {len(all_files)} files in normalized directory")

        # Print all files for debugging
        logger.info("Files in directory:")
        for file in all_files:
            logger.info(f"  - {file}")
    except Exception as e:
        logger.error(f"ERROR listing directory {normalized_path}: {e}")
        return dfs

    # Load each essential file
    for key, filename in essential_files.items():
        # Look for any file that ends with the essential filename
        matching_files = [f for f in all_files if f.endswith(filename)]

        if matching_files:
            # Use the first matching file
            file_path = os.path.join(normalized_path, matching_files[0])
            logger.info(f"Loading {key} data from: {matching_files[0]}")

            try:
                # Load the file
                df = pd.read_csv(file_path, low_memory=False)

                # Standardize ID column name
                if 'AvatarKey' in df.columns and 'PATIENT_ID' not in df.columns:
                    df = df.rename(columns={'AvatarKey': 'PATIENT_ID'})
                    logger.info(f"Renamed 'AvatarKey' to 'PATIENT_ID' in {key} data")

                # Store the dataframe
                dfs[key] = df
                logger.info(f"Successfully loaded {key} data: {df.shape[0]} records, {df.shape[1]} columns")
            except Exception as e:
                logger.error(f"ERROR loading {matching_files[0]}: {e}")
        else:
            logger.warning(f"WARNING: No file found ending with: {filename}")

    # Log summary of loaded data
    logger.info("==================== DATA LOADING SUMMARY ====================")
    logger.info(f"Successfully loaded {len(dfs)} clinical data files:")
    for key, df in dfs.items():
        if df is not None:
            logger.info(f"  - {key}: {df.shape[0]} records, {df.shape[1]} columns")
    
    return dfs

def _process_icb_data(medications):
    """
    Process medications data to identify ICB treatments.
    
    Args:
        medications (pd.DataFrame): Medications data
        
    Returns:
        pd.DataFrame: Processed ICB data with patient ID and earliest ICB treatment
    """
    logger.info("Processing ICB medications data")

    # Check if medications dataframe exists
    if medications is None or len(medications) == 0:
        logger.warning("No medications data available")
        return pd.DataFrame()

    # Print the actual columns in the medications dataframe for debugging
    logger.info(f"Available columns in medications data: {medications.columns.tolist()}")

    # Map standard column names to possible variations
    column_mappings = {
        'medication_name': ['TreatmentName', 'Medication', 'MedicationName', 'DrugName', 'Drug'],
        'start_age': ['TreatmentStartAge', 'StartAge', 'AgeAtMedStart', 'MedicationStartAge']
    }

    # Find the actual column names in the medications dataframe
    actual_columns = {}
    for standard_name, possible_names in column_mappings.items():
        for name in possible_names:
            if name in medications.columns:
                actual_columns[standard_name] = name
                logger.info(f"Using '{name}' for {standard_name}")
                break

    # Check if we found the required columns
    if 'medication_name' not in actual_columns:
        logger.warning("Could not find a medication name column in the medications data")
        return pd.DataFrame()

    # Make a copy of the medications dataframe
    meds_df = medications.copy()

    # Create a lowercase version of the medication name for matching
    medication_name_col = actual_columns['medication_name']
    meds_df['medication_lower'] = meds_df[medication_name_col].astype(str).str.lower()

    # Identify ICB drugs
    meds_df['is_icb'] = meds_df['medication_lower'].apply(
        lambda x: any(drug in x for drug in ICB_DRUGS_FLAT) if pd.notna(x) else False
    )

    # Filter to ICB treatments
    icb_meds = meds_df[meds_df['is_icb']].copy()

    if len(icb_meds) == 0:
        logger.info("No ICB treatments found in medications data")
        return pd.DataFrame()

    logger.info(f"Found {len(icb_meds)} ICB treatment records for {icb_meds['PATIENT_ID'].nunique()} patients")

    # Identify ICB class for each medication
    def get_icb_class(drug_name):
        drug_name_lower = str(drug_name).lower()
        for icb_class, drugs in ICB_DRUGS.items():
            if any(drug.lower() in drug_name_lower for drug in drugs):
                return icb_class
        return 'Unknown'

    icb_meds['ICB_class'] = icb_meds[medication_name_col].apply(get_icb_class)

    # Get the start age column if available
    start_age_col = actual_columns.get('start_age')

    # If start age column is available, use it for sorting to get earliest ICB
    if start_age_col:
        # Convert to numeric, coercing errors to NaN
        icb_meds[start_age_col] = pd.to_numeric(icb_meds[start_age_col], errors='coerce')
        # Sort by start age and get earliest for each patient
        icb_earliest = icb_meds.sort_values(start_age_col).groupby('PATIENT_ID').first().reset_index()
    else:
        # If no start age available, just get the first record for each patient
        logger.warning("No start age column found in medications data. Using first record for each patient.")
        icb_earliest = icb_meds.groupby('PATIENT_ID').first().reset_index()

    # Prepare output columns
    output_columns = {
        'PATIENT_ID': 'PATIENT_ID',
        medication_name_col: 'ICB_DRUG',
        'ICB_class': 'ICB_CLASS'
    }

    # Add start age column if available
    if start_age_col:
        output_columns[start_age_col] = 'ICB_START_AGE'

    # Select and rename columns
    icb_earliest_renamed = icb_earliest[list(output_columns.keys())].rename(columns=output_columns)

    # Add ICB status column
    icb_earliest_renamed['ICB_STATUS'] = 'YES'

    logger.info(f"Processed ICB data for {len(icb_earliest_renamed)} patients")

    return icb_earliest_renamed

def _add_diagnosis_info(patients, diagnosis):
    """
    Add diagnosis information to patient dataframe.
    
    Args:
        patients (pd.DataFrame): Patient data
        diagnosis (pd.DataFrame): Diagnosis data
        
    Returns:
        pd.DataFrame: Updated patient data with diagnosis information
    """
    logger.info("Adding diagnosis information")

    # Create a copy to avoid modifying the original
    patients_copy = patients.copy()

    # Check for required columns
    required_cols = ['PATIENT_ID', 'HistologyCode']

    for col in required_cols:
        if col not in diagnosis.columns:
            logger.warning(f"Missing required column in diagnosis data: {col}")
            if col == 'PATIENT_ID':
                return patients_copy

    # Get melanoma diagnoses based on histology code
    diagnosis['CancerType'] = diagnosis['HistologyCode'].apply(get_cancer_type)

    # Log the available columns in the diagnosis dataframe
    logger.info(f"Available columns in diagnosis data: {diagnosis.columns.tolist()}")

    # Check for a diagnosis age column - use the first available one
    age_cols = ['DiagnosisAge', 'AgeAtDiagnosis', 'Age', 'DiagnosisAgeInYears']
    diagnosis_age_col = None

    for col in age_cols:
        if col in diagnosis.columns:
            diagnosis_age_col = col
            logger.info(f"Using {col} as the diagnosis age column")
            break
                
    # If no diagnosis age column is found, sort by PATIENT_ID
    if diagnosis_age_col:
        # For each patient, get their earliest diagnosis record
        earliest_diagnosis = diagnosis.sort_values(diagnosis_age_col).groupby('PATIENT_ID').first().reset_index()
    else:
        logger.warning("No diagnosis age column found. Using the first diagnosis record for each patient.")
        earliest_diagnosis = diagnosis.groupby('PATIENT_ID').first().reset_index()

    # Select columns to merge into patient data
    diagnosis_cols = [
        'PATIENT_ID', 'CancerType', 'HistologyCode', 'HistologyDescription', 'PrimarySite',
        'TumorLocation', 'STAGE'
    ]

    # Add the diagnosis age column if it exists
    if diagnosis_age_col:
        diagnosis_cols.append(diagnosis_age_col)

    # Check for a diagnosis year column
    year_cols = ['DiagnosisYear', 'YearOfDiagnosis']
    diagnosis_year_col = None

    for col in year_cols:
        if col in diagnosis.columns:
            diagnosis_year_col = col
            diagnosis_cols.append(col)
            logger.info(f"Using {col} as the diagnosis year column")
            break
        
    # Keep only columns that exist
    diagnosis_cols = [col for col in diagnosis_cols if col in earliest_diagnosis.columns]
    # Merge diagnosis information with patient data
    patients_merged = patients_copy.merge(
        earliest_diagnosis[diagnosis_cols],
        on='PATIENT_ID',
        how='left'
    )

    # Rename the age column to DiagnosisAge if it's not already named that
    if diagnosis_age_col and diagnosis_age_col != 'DiagnosisAge':
        patients_merged = patients_merged.rename(columns={diagnosis_age_col: 'DiagnosisAge'})
        logger.info(f"Renamed {diagnosis_age_col} to DiagnosisAge")

    # Rename the year column to DiagnosisYear if it's not already named that
    if diagnosis_year_col and diagnosis_year_col != 'DiagnosisYear':
        patients_merged = patients_merged.rename(columns={diagnosis_year_col: 'DiagnosisYear'})
        logger.info(f"Renamed {diagnosis_year_col} to DiagnosisYear")

    # Clean stage values
    if 'STAGE' in patients_merged.columns:
        patients_merged['STAGE_SIMPLE'] = patients_merged['STAGE'].apply(clean_stage_value)

    logger.info(f"Added diagnosis info to {len(patients_merged)} patients")

    return patients_merged

def _add_survival_info(patients, vital):
    """
    Add vital status and survival information to patient dataframe.
    
    Args:
        patients (pd.DataFrame): Patient data
        vital (pd.DataFrame): Vital status data
    
    Returns:
        pd.DataFrame: Updated patient data with survival information
    """
    logger.info("Adding survival information")

    # Create a copy to avoid modifying the original
    patients_copy = patients.copy()

    # Log the available columns in the vital dataframe
    logger.info(f"Available columns in vital data: {vital.columns.tolist()}")

    # Check for a patient ID column
    if 'PATIENT_ID' not in vital.columns:
        logger.warning("No PATIENT_ID column found in vital data")
        return patients_copy

    # Check for a vital status column
    vital_status_cols = ['VitalStatus', 'Status', 'PatientStatus', 'DeathInd']
    vital_status_col = None

    for col in vital_status_cols:
        if col in vital.columns:
            vital_status_col = col
            logger.info(f"Using {col} as the vital status column")
            break

    # Check for a last contact age column
    last_contact_cols = ['LastContactAge', 'AgeAtLastContact', 'LastFollowUpAge', 'FollowUpAge']
    last_contact_col = None

    for col in last_contact_cols:
        if col in vital.columns:
            last_contact_col = col
            logger.info(f"Using {col} as the last contact age column")
            break

    # Check for a last contact year column
    last_contact_year_cols = ['LastContactYear', 'YearOfLastContact', 'LastFollowUpYear']
    last_contact_year_col = None

    for col in last_contact_year_cols:
        if col in vital.columns:
            last_contact_year_col = col
            logger.info(f"Using {col} as the last contact year column")
            break

    # Check for a death cause column
    death_cause_cols = ['DeathCause', 'CauseOfDeath', 'DeathDetails']
    death_cause_col = None

    for col in death_cause_cols:
        if col in vital.columns:
            death_cause_col = col
            logger.info(f"Using {col} as the death cause column")
            break

    # Build a list of columns to select from the vital dataframe
    vital_cols = ['PATIENT_ID']

    if vital_status_col:
        vital_cols.append(vital_status_col)

    if last_contact_col:
        vital_cols.append(last_contact_col)

    if last_contact_year_col:
        vital_cols.append(last_contact_year_col)

    if death_cause_col:
        vital_cols.append(death_cause_col)

    # Keep only columns that exist
    vital_cols = [col for col in vital_cols if col in vital.columns]

    if len(vital_cols) <= 1:
        logger.warning("No useful vital status columns found after filtering")
        return patients_copy

    # Merge vital status information with patient data
    patients_merged = patients_copy.merge(
        vital[vital_cols],
        on='PATIENT_ID',
        how='left'
    )

    # Create OS_STATUS (1 for deceased, 0 for alive)
    if vital_status_col:
        # Map various status values to binary (1=deceased, 0=alive/censored)
        status_mapping = {
            'DECEASED': 1, 'DEAD': 1, 'DIED': 1, 'DEATH': 1, 'EXPIRED': 1, 'YES': 1, '1': 1, 'TRUE': 1,
            'LIVING': 0, 'ALIVE': 0, 'NO': 0, '0': 0, 'FALSE': 0, 'CENSORED': 0
        }

        # Apply the mapping and convert to uppercase to standardize
        patients_merged['OS_STATUS'] = patients_merged[vital_status_col].astype(str).str.upper().map(
            lambda x: status_mapping.get(x, 0)  # Default to 0 (alive) if unknown
        )

        # Rename the vital status column to VitalStatus if it's not already named that
        if vital_status_col != 'VitalStatus':
            patients_merged = patients_merged.rename(columns={vital_status_col: 'VitalStatus'})
            logger.info(f"Renamed {vital_status_col} to VitalStatus")
    else:
        logger.warning("No vital status column found. Assuming all patients are alive.")
        patients_merged['OS_STATUS'] = 0

    # Rename the last contact columns to standardized names
    if last_contact_col and last_contact_col != 'LastContactAge':
        patients_merged = patients_merged.rename(columns={last_contact_col: 'LastContactAge'})
        logger.info(f"Renamed {last_contact_col} to LastContactAge")

    if last_contact_year_col and last_contact_year_col != 'LastContactYear':
        patients_merged = patients_merged.rename(columns={last_contact_year_col: 'LastContactYear'})
        logger.info(f"Renamed {last_contact_year_col} to LastContactYear")

    if death_cause_col and death_cause_col != 'DeathCause':
        patients_merged = patients_merged.rename(columns={death_cause_col: 'DeathCause'})
        logger.info(f"Renamed {death_cause_col} to DeathCause")

    # Calculate survival time if diagnosis age and last contact age are available
    if 'DiagnosisAge' in patients_merged.columns and 'LastContactAge' in patients_merged.columns:
        # First ensure both columns are numeric
        for col in ['DiagnosisAge', 'LastContactAge']:
            try:
                patients_merged[col] = pd.to_numeric(patients_merged[col], errors='coerce')
            except Exception as e:
                logger.warning(f"Could not convert {col} to numeric: {e}")

        # Calculate OS_TIME (in months)
        patients_merged['OS_TIME'] = patients_merged['LastContactAge'] - patients_merged['DiagnosisAge']

        # Convert to months (assumes ages are in years)
        patients_merged['OS_TIME'] = patients_merged['OS_TIME'] * 12

        # Remove negative survival times
        negative_time = (patients_merged['OS_TIME'] < 0)
        if negative_time.any():
            logger.warning(f"Found {negative_time.sum()} patients with negative survival time. Setting to NaN.")
            patients_merged.loc[negative_time, 'OS_TIME'] = np.nan

        logger.info(f"Calculated survival time (OS_TIME) for {patients_merged['OS_TIME'].notna().sum()} patients")
    else:
        logger.warning("Cannot calculate survival time: missing either DiagnosisAge or LastContactAge")

    logger.info(f"Added survival info to {len(patients_merged)} patients")
    return patients_merged

def _add_surgery_info(clinical, surgery_df):
    """
    Add surgery information to the clinical dataframe.
    
    Args:
        clinical: Base clinical dataframe
        surgery_df: Surgery dataframe
        
    Returns:
        DataFrame with added surgery information
    """
    try:
        logger.info("Processing surgery information")
        
        # Make a copy of the surgery dataframe
        surgery = surgery_df.copy()
        
        # Ensure surgery data has PATIENT_ID column
        if 'PATIENT_ID' not in surgery.columns and 'AvatarKey' in surgery.columns:
            surgery = surgery.rename(columns={'AvatarKey': 'PATIENT_ID'})
            logger.info("Renamed AvatarKey to PATIENT_ID in surgery data")
        elif 'PATIENT_ID' not in surgery.columns:
            # Try to find a patient ID column
            potential_id_cols = [col for col in surgery.columns if 'patient' in col.lower() or 'avatar' in col.lower() or 'key' in col.lower()]
            if potential_id_cols:
                surgery = surgery.rename(columns={potential_id_cols[0]: 'PATIENT_ID'})
                logger.info(f"Renamed {potential_id_cols[0]} to PATIENT_ID in surgery data")
            else:
                logger.error("No patient ID column found in surgery data")
                return clinical  # Return without modification
        
        # Count surgeries per patient
        surgery_counts = surgery.groupby('PATIENT_ID').size().reset_index(name='SurgeryCount')
        
        # Add surgery count to clinical data
        clinical = clinical.merge(surgery_counts, on='PATIENT_ID', how='left')
        
        # Fill missing values with 0 (patients with no surgeries)
        clinical['SurgeryCount'] = clinical['SurgeryCount'].fillna(0).astype(int)
        
        # Add flag for patients who had surgery
        clinical['HadSurgery'] = (clinical['SurgeryCount'] > 0).astype(int)
        
        logger.info(f"Added surgery information for {len(surgery_counts)} patients")
        
        return clinical
        
    except Exception as e:
        logger.error(f"Error adding surgery information: {e}")
        logger.error(traceback.format_exc())
        return clinical

def integrate_molecular_data(clinical, base_path, dfs=None, diagnosis_df=None):
    """
    Save the final processed clinical data.
    Previously integrated molecular data (TMB/MSI), but now simplified to just save the final data
    after melanoma filtering.
    
    Args:
        clinical: Clinical data DataFrame (already filtered for melanoma patients)
        base_path: Base path for the project
        dfs: Dictionary of dataframes loaded by load_clinical_data
        diagnosis_df: Diagnosis dataframe for additional verification of melanoma samples
        
    Returns:
        DataFrame with the final processed clinical data
    """
    try:
        # Create a copy of the clinical data
        clinical_with_molecular = clinical.copy()
        
        # Get the list of melanoma patient IDs from the clinical data
        melanoma_patient_ids = clinical_with_molecular['PATIENT_ID'].unique()
        logger.info(f"Working with {len(melanoma_patient_ids)} melanoma patients for final data preparation")

        # Get tumor sequencing data for informational purposes only
        if dfs is not None and "tumor_sequencing" in dfs:
            tumor_seq = dfs["tumor_sequencing"].copy()
            
            # Ensure we have a PATIENT_ID column
            if 'PATIENT_ID' not in tumor_seq.columns and 'AvatarKey' in tumor_seq.columns:
                tumor_seq = tumor_seq.rename(columns={'AvatarKey': 'PATIENT_ID'})
            
            # Filter tumor sequencing data for melanoma patients
            tumor_seq_melanoma = tumor_seq[tumor_seq['PATIENT_ID'].isin(melanoma_patient_ids)].copy()
            logger.info(f"Filtered tumor sequencing data for melanoma patients: {len(tumor_seq_melanoma)} out of {len(tumor_seq)} total records")

            # Link tumor sequencing to diagnosis data if available (just for reporting)
            if diagnosis_df is not None:
                try:
                    # Ensure diagnosis_df has PATIENT_ID column
                    if 'PATIENT_ID' not in diagnosis_df.columns and 'AvatarKey' in diagnosis_df.columns:
                        diagnosis_df = diagnosis_df.rename(columns={'AvatarKey': 'PATIENT_ID'})

                    logger.info(f"Attempting to link tumor sequencing to diagnosis data")

                    # Check for columns needed to link tumor sequencing to specific diagnoses
                    link_columns = {
                        'tumor_seq': ['SpecimenID', 'SampleCollectionAge', 'SampleID', 'SLID', 'DiagnosisID'],
                        'diagnosis': ['DiagnosisID', 'AgeAtDiagnosis', 'HistologyCode', 'HISTOLOGY_CODE']
                    }

                    # Check which columns exist
                    tumor_link_cols = [col for col in link_columns['tumor_seq'] if col in tumor_seq_melanoma.columns]
                    diag_link_cols = [col for col in link_columns['diagnosis'] if col in diagnosis_df.columns]

                    if tumor_link_cols and diag_link_cols:
                        logger.info(f"Found columns to attempt linking: Tumor: {tumor_link_cols}, Diagnosis: {diag_link_cols}")
                except Exception as e:
                    logger.error(f"Error linking tumor sequencing to diagnosis: {e}")
                    logger.error(traceback.format_exc())

        # Save the final processed data
        output_dir = os.path.join(base_path, "processed_data")
        os.makedirs(output_dir, exist_ok=True)
        
        # Add a note that this is melanoma-only data
        clinical_with_molecular['data_type'] = 'melanoma_only'

        # Save as CSV
        clinical_with_molecular.to_csv(os.path.join(output_dir, "processed_clinical_molecular.csv"), index=False)
        logger.info(f"Saved final processed clinical data for {len(clinical_with_molecular)} melanoma patients")

        # Try to save as Parquet if libraries are available
        try:
            # Check if either pyarrow or fastparquet is available
            pyarrow_spec = importlib.util.find_spec("pyarrow")
            fastparquet_spec = importlib.util.find_spec("fastparquet")

            if pyarrow_spec is not None or fastparquet_spec is not None:
                clinical_with_molecular.to_parquet(os.path.join(output_dir, "processed_clinical_molecular.parquet"), index=False)
                logger.info(f"Saved processed clinical data in Parquet format")
            else:
                logger.warning("Skipping Parquet export: neither pyarrow nor fastparquet is installed. Install one of these packages for Parquet support.")
        except Exception as e:
            logger.error(f"Error saving parquet file: {e}")
        
        return clinical_with_molecular
        
    except Exception as e:
        logger.error(f"Error preparing final clinical data: {e}")
        logger.error(traceback.format_exc())
        return clinical

def save_processed_data(df, output_dir):
    """
    Save processed data to CSV and Parquet formats
    
    Args:
        df (pandas.DataFrame): Dataframe to save
        output_dir (str): Directory to save the data
    """
    try:
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        # Save as CSV
        csv_file = os.path.join(output_dir, 'processed_clinical.csv')
        df.to_csv(csv_file, index=False)
        logger.info(f"Saved processed data to {csv_file}")
        
        # Try to save as Parquet (if dependencies are available)
        try:
            # Check if either pyarrow or fastparquet is available
            pyarrow_spec = importlib.util.find_spec("pyarrow")
            fastparquet_spec = importlib.util.find_spec("fastparquet")
            
            if pyarrow_spec is not None or fastparquet_spec is not None:
                parquet_file = os.path.join(output_dir, 'processed_clinical.parquet')
                df.to_parquet(parquet_file)
                logger.info(f"Saved processed data to {parquet_file}")
            else:
                logger.warning("Skipping Parquet export: neither pyarrow nor fastparquet is installed. "
                              "Install one of these packages for Parquet support.")
        except Exception as e:
            logger.warning(f"Skipping Parquet export: {str(e)}")
            
    except Exception as e:
        logger.error(f"Error saving processed data: {str(e)}")

def setup_output_dirs(base_path):
    """
    Set up output directories for plots and data.
    """
    # Check if base_path already ends with 'codes'
    if base_path.endswith('codes'):
        output_dir = os.path.join(base_path, "output", "data_analysis")
    else:
        output_dir = os.path.join(base_path, "codes", "output", "data_analysis")
    
    # Create the output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Print the absolute path for debugging
    print(f"Setup output directory at: {os.path.abspath(output_dir)}")
    
    return output_dir

def save_plot(fig, filename, output_dir):
    """
    Save a matplotlib figure to a file.
    
    Args:
        fig: matplotlib figure object
        filename: name of the file to save
        output_dir: directory to save the file in
    """
    # Fix for duplicate extensions - remove any existing .png extension before adding one
    if filename.endswith('.png'):
        filename = filename[:-4]  # Remove .png extension
    
    # Create full path
    full_path = os.path.join(output_dir, f"{filename}.png")
    
    # Save the figure
    fig.savefig(full_path, bbox_inches='tight', dpi=300)
    
    # Print confirmation message with full path
    print(f"Saved plot to: {os.path.abspath(full_path)}")
    
    # Close the figure to release memory
    plt.close(fig)
    
    return full_path

def plot_demographics(df, output_dir=None):
    """
    Create demographic plots for the clinical cohort.
    
    Args:
        df (pd.DataFrame): Clinical data
        output_dir (str, optional): Directory to save plots

    Returns:
        str or None: Path to demographics directory if plots were saved, otherwise None
    """
    logger.info("Creating demographic plots")

    try:
        # Create demographics directory if specified
        if output_dir:
            demographics_dir = os.path.join(output_dir, 'demographics')
            os.makedirs(demographics_dir, exist_ok=True)
        else:
            demographics_dir = None

        # 1. Age Distribution
        if 'DiagnosisAge' in df.columns and df['DiagnosisAge'].notna().any():
            try:
                valid_age = df[df['DiagnosisAge'].notna()].copy()

                plt.figure(figsize=(10, 6))
                sns.histplot(valid_age['DiagnosisAge'], bins=20, kde=True)
                plt.xlabel('Age at Diagnosis (years)')
                plt.ylabel('Count')
                plt.title(f'Age Distribution at Diagnosis (n={len(valid_age)})')
                plt.grid(alpha=0.3)

                if demographics_dir:
                    plt.savefig(os.path.join(demographics_dir, 'age_distribution.png'),
                                bbox_inches='tight', dpi=300)
                    plt.close()
                    logger.info(f"Saved age distribution plot with {len(valid_age)} patients")
            except Exception as e:
                logger.error(f"Error creating age distribution plot: {e}")
        else:
            logger.warning("Cannot create age distribution: missing DiagnosisAge column or no valid values")

        # 2. Sex Distribution
        if 'Sex' in df.columns and df['Sex'].notna().any():
            try:
                # Count by sex
                sex_count = df['Sex'].value_counts()

                plt.figure(figsize=(8, 6))
                sex_count.plot(kind='bar', color='steelblue')
                plt.xlabel('Sex')
                plt.ylabel('Count')
                plt.title(f'Sex Distribution (n={len(df)})')
                plt.grid(alpha=0.3, axis='y')

                # Add count and percentage labels
                for i, count in enumerate(sex_count):
                    percentage = 100 * count / sex_count.sum()
                    plt.text(i, count + 1, f"{count} ({percentage:.1f}%)",
                             ha='center', va='bottom', fontsize=11)

                if demographics_dir:
                    plt.savefig(os.path.join(demographics_dir, 'sex_distribution.png'),
                                bbox_inches='tight', dpi=300)
                    plt.close()
                    logger.info(f"Saved sex distribution plot with {len(df)} patients")
            except Exception as e:
                logger.error(f"Error creating sex distribution plot: {e}")
        else:
            logger.warning("Cannot create sex distribution: missing Sex column or no valid values")

        # 3. Race Distribution
        if 'Race' in df.columns and df['Race'].notna().any():
            try:
                # Count by race
                race_count = df['Race'].value_counts()

                # If too many categories, keep only top 5 plus "Other"
                if len(race_count) > 5:
                    top_races = race_count.nlargest(5).index.tolist()
                    df_race = df.copy()
                    df_race['Race_Group'] = df_race['Race'].apply(lambda x: x if x in top_races else 'Other')
                    race_count = df_race['Race_Group'].value_counts()

                plt.figure(figsize=(10, 6))
                race_count.plot(kind='bar', color='steelblue')
                plt.xlabel('Race')
                plt.ylabel('Count')
                plt.title(f'Race Distribution (n={len(df)})')
                plt.grid(alpha=0.3, axis='y')
                plt.xticks(rotation=45, ha='right')

                if demographics_dir:
                    plt.savefig(os.path.join(demographics_dir, 'race_distribution.png'),
                                bbox_inches='tight', dpi=300)
                    plt.close()
                    logger.info(f"Saved race distribution plot with {len(df)} patients")
            except Exception as e:
                logger.error(f"Error creating race distribution plot: {e}")
        else:
            logger.warning("Cannot create race distribution: missing Race column or no valid values")

        # 4. Cancer Type Distribution (if mixed cohort)
        if 'CancerType' in df.columns and df['CancerType'].nunique() > 1:
            try:
                # Count by cancer type
                cancer_count = df['CancerType'].value_counts()

                plt.figure(figsize=(10, 6))
                cancer_count.plot(kind='bar', color='steelblue')
                plt.xlabel('Cancer Type')
                plt.ylabel('Count')
                plt.title(f'Cancer Type Distribution (n={len(df)})')
                plt.grid(alpha=0.3, axis='y')

                # Add count and percentage labels
                for i, count in enumerate(cancer_count):
                    percentage = 100 * count / cancer_count.sum()
                    plt.text(i, count + 1, f"{count} ({percentage:.1f}%)",
                             ha='center', va='bottom', fontsize=11)

                if demographics_dir:
                    plt.savefig(os.path.join(demographics_dir, 'cancer_type_distribution.png'),
                                bbox_inches='tight', dpi=300)
                    plt.close()
                    logger.info(f"Saved cancer type distribution plot with {len(df)} patients")
            except Exception as e:
                logger.error(f"Error creating cancer type distribution plot: {e}")

        # 5. ICB Status Distribution (if available)
        if 'ICB_STATUS' in df.columns or 'HAS_ICB' in df.columns:
            try:
                # Determine which ICB column to use
                icb_col = 'ICB_STATUS' if 'ICB_STATUS' in df.columns else 'HAS_ICB'

                # Fill missing values as 'No'
                df_icb = df.copy()
                if icb_col == 'ICB_STATUS':
                    df_icb[icb_col] = df_icb[icb_col].fillna('NO')
                else:  # HAS_ICB is likely numeric (0/1)
                    df_icb[icb_col] = df_icb[icb_col].fillna(0)

                # Count by ICB status
                icb_count = df_icb[icb_col].value_counts()

                plt.figure(figsize=(8, 6))
                icb_count.plot(kind='bar', color='steelblue')
                plt.xlabel('ICB Treatment Status')
                plt.ylabel('Count')
                plt.title(f'ICB Treatment Distribution (n={len(df_icb)})')
                plt.grid(alpha=0.3, axis='y')

                # Add count and percentage labels
                for i, count in enumerate(icb_count):
                    percentage = 100 * count / icb_count.sum()
                    plt.text(i, count + 1, f"{count} ({percentage:.1f}%)",
                             ha='center', va='bottom', fontsize=11)

                if demographics_dir:
                    plt.savefig(os.path.join(demographics_dir, 'icb_distribution.png'),
                                bbox_inches='tight', dpi=300)
                    plt.close()
                    logger.info(f"Saved ICB distribution plot with {len(df_icb)} patients")
            except Exception as e:
                logger.error(f"Error creating ICB distribution plot: {e}")
    
        return demographics_dir

    except Exception as e:
        logger.error(f"Error in plot_demographics: {e}")
        logger.error(traceback.format_exc())
        return None

def plot_molecular_features(df, output_dir):
    """
    Create plots for molecular features.
    
    Args:
        df: pandas DataFrame with molecular data
        output_dir: directory to save the plots
    """
    # Create molecular directory
    molecular_dir = os.path.join(output_dir, "molecular")
    os.makedirs(molecular_dir, exist_ok=True)
    print(f"Created molecular directory at: {os.path.abspath(molecular_dir)}")
    
    # Check for molecular features
    molecular_features = ['TMB', 'MSI', 'CD8_SCORE']
    available_features = [col for col in molecular_features if col in df.columns]
    
    if not available_features:
        print("Warning: No molecular features found in the dataset")
        # Create an empty placeholder file to indicate processing was attempted
        with open(os.path.join(molecular_dir, "NO_MOLECULAR_DATA.txt"), "w") as f:
            f.write("No molecular data was found in the dataset. The following columns were checked: " + 
                    ", ".join(molecular_features) + "\n\n" +
                    "This file was created on " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        return molecular_dir
    
    print(f"Found molecular features: {available_features}")
    
    # TMB distribution if available
    if 'TMB' in df.columns:
        # Convert to numeric, coercing errors to NaN
        df['TMB'] = pd.to_numeric(df['TMB'], errors='coerce')
        valid_tmb = df['TMB'].dropna()
        
        if len(valid_tmb) > 0:
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.hist(valid_tmb, bins=20, alpha=0.7, color='skyblue', edgecolor='black')
            ax.set_xlabel('Tumor Mutational Burden (mutations/Mb)', fontsize=12)
            ax.set_ylabel('Number of Patients', fontsize=12)
            ax.set_title(f'TMB Distribution (n={len(valid_tmb)})', fontsize=14)
            ax.grid(alpha=0.3)
            
            # Add mean and median lines
            mean_tmb = valid_tmb.mean()
            median_tmb = valid_tmb.median()
            ax.axvline(mean_tmb, color='red', linestyle='--', linewidth=1, 
                      label=f'Mean: {mean_tmb:.2f}')
            ax.axvline(median_tmb, color='green', linestyle='--', linewidth=1,
                      label=f'Median: {median_tmb:.2f}')
            ax.legend()
            
            # Save the plot
            save_plot(fig, "tmb_distribution", molecular_dir)
            print(f"Created TMB distribution plot with {len(valid_tmb)} patients")
        else:
            print("Warning: No valid TMB data found")
            with open(os.path.join(molecular_dir, "NO_VALID_TMB_DATA.txt"), "w") as f:
                f.write("TMB column exists but contains no valid numeric data")
    
    # MSI distribution if available
    if 'MSI' in df.columns:
        # Convert to numeric, coercing errors to NaN
        df['MSI'] = pd.to_numeric(df['MSI'], errors='coerce')
        valid_msi = df['MSI'].dropna()
        
        if len(valid_msi) > 0:
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.hist(valid_msi, bins=20, alpha=0.7, color='lightgreen', edgecolor='black')
            ax.set_xlabel('Microsatellite Instability Score', fontsize=12)
            ax.set_ylabel('Number of Patients', fontsize=12)
            ax.set_title(f'MSI Distribution (n={len(valid_msi)})', fontsize=14)
            ax.grid(alpha=0.3)
            
            # Save the plot
            save_plot(fig, "msi_distribution", molecular_dir)
            print(f"Created MSI distribution plot with {len(valid_msi)} patients")
        else:
            print("Warning: No valid MSI data found")
            with open(os.path.join(molecular_dir, "NO_VALID_MSI_DATA.txt"), "w") as f:
                f.write("MSI column exists but contains no valid numeric data\n" +
                        "File created: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    
    # CD8 score distribution if available
    if 'CD8_SCORE' in df.columns:
        # Convert to numeric, coercing errors to NaN
        df['CD8_SCORE'] = pd.to_numeric(df['CD8_SCORE'], errors='coerce')
        valid_cd8 = df['CD8_SCORE'].dropna()
        
        if len(valid_cd8) > 0:
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.hist(valid_cd8, bins=20, alpha=0.7, color='salmon', edgecolor='black')
            ax.set_xlabel('CD8 T-cell Score', fontsize=12)
            ax.set_ylabel('Number of Patients', fontsize=12)
            ax.set_title(f'CD8 Score Distribution (n={len(valid_cd8)})', fontsize=14)
            ax.grid(alpha=0.3)
            
            # Save the plot
            save_plot(fig, "cd8_distribution", molecular_dir)
            print(f"Created CD8 score distribution plot with {len(valid_cd8)} patients")
        else:
            print("Warning: No valid CD8 score data found")
            with open(os.path.join(molecular_dir, "NO_VALID_CD8_DATA.txt"), "w") as f:
                f.write("CD8_SCORE column exists but contains no valid numeric data\n" +
                        "File created: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    
    # If no plots were created but the function got this far, create a marker file
    if not any(f.endswith('.png') for f in os.listdir(molecular_dir)):
        with open(os.path.join(molecular_dir, "NO_VALID_MOLECULAR_DATA.txt"), "w") as f:
            f.write(f"Found columns {available_features} but could not create any plots due to invalid data\n" +
                    "File created: " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    
    return molecular_dir

def plot_survival_curves(df, output_dir):
    """
    Create Kaplan-Meier survival curves for the cohort.
    
    Args:
        df: pandas DataFrame with survival data
        output_dir: directory to save the plots
    """
    logger.info("Creating Kaplan-Meier survival curves")

    try:
        # Create survival plots directory
        survival_dir = os.path.join(output_dir, "survival")
        os.makedirs(survival_dir, exist_ok=True)

        # Check for required columns
        required_cols = ['OS_TIME', 'OS_EVENT']
        if not all(col in df.columns for col in required_cols):
            logger.warning("Required columns OS_TIME and OS_EVENT not found in dataset")
            return

        # Prepare data
        data = df.copy()
        data['OS_TIME'] = pd.to_numeric(data['OS_TIME'], errors='coerce')
        data['OS_EVENT'] = pd.to_numeric(data['OS_EVENT'], errors='coerce')

        # Drop rows with missing survival data
        data = data.dropna(subset=['OS_TIME', 'OS_EVENT'])

        if len(data) == 0:
            logger.warning("No valid survival data available after cleaning")
            return

        # Overall survival curve
        try:
            kmf = KaplanMeierFitter()
            kmf.fit(data['OS_TIME'], event_observed=data['OS_EVENT'])
            
            plt.figure(figsize=(10, 6))
            kmf.plot_survival_function(show_censors=True, ci_show=True)
            plt.title('Overall Survival')
            plt.xlabel('Time (months)')
            plt.ylabel('Survival Probability')
            plt.grid(alpha=0.3)
            
            save_plot(plt.gcf(), 'overall_survival', survival_dir)
            plt.close()
            
            logger.info(f"Created overall survival curve with {len(data)} patients")
        except Exception as e:
            logger.error(f"Error creating overall survival curve: {str(e)}")

        # Stratified by ICB treatment if available
        if 'ICB_STATUS' in data.columns or 'HAS_ICB' in data.columns:
            try:
                # Determine which ICB column to use
                icb_col = 'ICB_STATUS' if 'ICB_STATUS' in data.columns else 'HAS_ICB'
                
                # Create masks for ICB and non-ICB patients
                if icb_col == 'ICB_STATUS':
                    icb_mask = data[icb_col] == 'YES'
                else:
                    icb_mask = data[icb_col] == 1
                
                non_icb_mask = ~icb_mask

                if sum(icb_mask) > 0 or sum(non_icb_mask) > 0:
                    plt.figure(figsize=(10, 6))
                    
                    # Plot ICB group if it exists
                    if sum(icb_mask) > 0:
                        kmf_icb = KaplanMeierFitter()
                        kmf_icb.fit(data[icb_mask]['OS_TIME'], data[icb_mask]['OS_EVENT'], 
                                  label=f'ICB (n={sum(icb_mask)})')
                        kmf_icb.plot_survival_function(show_censors=True, ci_show=True)
                    
                    # Plot non-ICB group if it exists
                    if sum(non_icb_mask) > 0:
                        kmf_no_icb = KaplanMeierFitter()
                        kmf_no_icb.fit(data[non_icb_mask]['OS_TIME'], data[non_icb_mask]['OS_EVENT'], 
                                     label=f'No ICB (n={sum(non_icb_mask)})')
                        kmf_no_icb.plot_survival_function(show_censors=True, ci_show=True)
                    
                    plt.title('Overall Survival by ICB Treatment')
                    plt.xlabel('Time (months)')
                    plt.ylabel('Survival Probability')
                    plt.grid(alpha=0.3)
                    
                    save_plot(plt.gcf(), 'km_by_icb', survival_dir)
                    plt.close()
                    
                    # Statistical comparison if both groups exist
                    if sum(icb_mask) > 0 and sum(non_icb_mask) > 0:
                        try:
                            results = logrank_test(data[icb_mask]['OS_TIME'], data[non_icb_mask]['OS_TIME'],
                                                 data[icb_mask]['OS_EVENT'], data[non_icb_mask]['OS_EVENT'])
                            print(f"Log-rank test p-value (ICB vs No ICB): {results.p_value:.4f}")
                        except Exception as e:
                            logger.error(f"Error performing log-rank test: {str(e)}")
            except Exception as e:
                logger.error(f"Error creating ICB stratified survival curves: {str(e)}")
                
        # By sex if available
        if 'SEX' in data.columns:
            try:
                # Remove any NaN values in SEX column
                valid_sex = data[data['SEX'].notna()].copy()
                
                if len(valid_sex) > 0:
                    plt.figure(figsize=(10, 6))
                    
                    # Get unique sex values
                    sex_values = valid_sex['SEX'].unique()
                    
                    for sex in sex_values:
                        if pd.notna(sex) and sex != '':
                            mask = valid_sex['SEX'] == sex
                            if sum(mask) > 0:
                                kmf_sex = KaplanMeierFitter()
                                kmf_sex.fit(valid_sex[mask]['OS_TIME'], valid_sex[mask]['OS_EVENT'], label=f'Sex: {sex}')
                                kmf_sex.plot_survival_function()
                    
                    plt.xlabel('Time since diagnosis (years)')
                    plt.ylabel('Survival Probability')
                    plt.title('Overall Survival by Sex')
                    plt.grid(alpha=0.3)
                    save_plot(plt.gcf(), 'km_by_sex', survival_dir)
                    plt.close()
                    
                    # Statistical comparison if multiple sex categories exist
                    if len(sex_values) > 1:
                        try:
                            # Basic log-rank test for first two categories if they exist
                            if len(sex_values) >= 2 and pd.notna(sex_values[0]) and pd.notna(sex_values[1]):
                                mask1 = valid_sex['SEX'] == sex_values[0]
                                mask2 = valid_sex['SEX'] == sex_values[1]
                                
                                if sum(mask1) > 0 and sum(mask2) > 0:
                                    results = logrank_test(valid_sex[mask1]['OS_TIME'], valid_sex[mask2]['OS_TIME'],
                                                         valid_sex[mask1]['OS_EVENT'], valid_sex[mask2]['OS_EVENT'])
                                    print(f"Log-rank test p-value ({sex_values[0]} vs {sex_values[1]}): {results.p_value:.4f}")
                        except Exception as e:
                            logger.error(f"Error performing log-rank test for sex comparison: {str(e)}")
            except Exception as e:
                logger.error(f"Error creating sex-stratified survival curves: {str(e)}")
        
    except Exception as e:
        logger.error(f"Error analyzing survival: {str(e)}")
        logger.error(traceback.format_exc())

def create_stage_simple(clinical_df):
    """
    Create a simplified stage column (STAGE_SIMPLE) that tries to infer stage 
    from other available data for patients with Unknown stage.
    
    Parameters:
    -----------
    clinical_df : DataFrame
        Clinical data with stage information
        
    Returns:
    --------
    DataFrame
        Clinical data with new STAGE_SIMPLE column
    """
    logger.info("Creating simplified stage classification")
    
    # Make a copy to avoid modifying the original dataframe
    df = clinical_df.copy()
    
    # Start with existing stage (either Stage or StageAtICB) if available
    if 'StageAtICB' in df.columns:
        df['STAGE_SIMPLE'] = df['StageAtICB']
    elif 'Stage' in df.columns:
        df['STAGE_SIMPLE'] = df['Stage'].apply(clean_stage_value)
    else:
        df['STAGE_SIMPLE'] = 'Unknown'
    
    # Count initial unknown stages
    initial_unknown = (df['STAGE_SIMPLE'] == 'Unknown').sum()
    logger.info(f"Initial Unknown stages: {initial_unknown} ({initial_unknown/len(df):.1%} of patients)")
    
    # Try to infer stage from other fields for those with Unknown stage
    unknown_mask = df['STAGE_SIMPLE'] == 'Unknown'
    
    # Log fields we'll use for inference
    inference_fields = []
    for field in ['MetastaticDisease', 'MetastaticSites', 'DistantMetastasis', 'T', 'N', 'M']:
        if field in df.columns:
            inference_fields.append(field)
    
    logger.info(f"Using the following fields for stage inference: {inference_fields}")
    
    # Check for metastatic disease indicators
    if 'MetastaticDisease' in df.columns:
        # If metastatic disease is present, set stage to IV
        metastatic_mask = unknown_mask & (df['MetastaticDisease'].str.contains('Yes|Positive|Present', case=False, na=False))
        df.loc[metastatic_mask, 'STAGE_SIMPLE'] = 'IV'
        logger.info(f"Inferred stage IV for {metastatic_mask.sum()} patients based on MetastaticDisease")
    
    if 'MetastaticSites' in df.columns:
        # If metastatic sites are listed, set stage to IV
        metastatic_mask = unknown_mask & df['MetastaticSites'].notna() & (df['MetastaticSites'] != '') & (df['MetastaticSites'] != 'Unknown')
        df.loc[metastatic_mask, 'STAGE_SIMPLE'] = 'IV'
        logger.info(f"Inferred stage IV for {metastatic_mask.sum()} patients based on MetastaticSites")
    
    if 'M' in df.columns:
        # If M1, set stage to IV
        metastatic_mask = unknown_mask & df['M'].str.contains('M1', na=False)
        df.loc[metastatic_mask, 'STAGE_SIMPLE'] = 'IV'
        logger.info(f"Inferred stage IV for {metastatic_mask.sum()} patients based on M classification")
    
    # Use T and N classifications to infer earlier stages
    if all(col in df.columns for col in ['T', 'N']):
        # T1-2, N0, assume stage I
        stage1_mask = unknown_mask & df['T'].str.contains('T1|T2', na=False) & df['N'].str.contains('N0', na=False)
        df.loc[stage1_mask, 'STAGE_SIMPLE'] = 'I'
        logger.info(f"Inferred stage I for {stage1_mask.sum()} patients based on T1-2, N0")
        
        # T3-4, N0, assume stage II
        stage2_mask = unknown_mask & df['T'].str.contains('T3|T4', na=False) & df['N'].str.contains('N0', na=False)
        df.loc[stage2_mask, 'STAGE_SIMPLE'] = 'II'
        logger.info(f"Inferred stage II for {stage2_mask.sum()} patients based on T3-4, N0")
        
        # Any T, N1-3, M0, assume stage III
        stage3_mask = unknown_mask & df['N'].str.contains('N1|N2|N3', na=False) & ~df['T'].str.contains('Tx', na=False)
        if 'M' in df.columns:
            stage3_mask = stage3_mask & (df['M'].str.contains('M0', na=False) | df['M'].isna())
        df.loc[stage3_mask, 'STAGE_SIMPLE'] = 'III'
        logger.info(f"Inferred stage III for {stage3_mask.sum()} patients based on N1-3")
    
    # Additional logic for melanoma-specific staging
    if 'Histology' in df.columns:
        melanoma_mask = df['Histology'].str.contains('melanoma', case=False, na=False)
        logger.info(f"Found {melanoma_mask.sum()} melanoma patients")
        
        # For melanoma with lymph node involvement, assume at least stage III
        if 'N' in df.columns:
            mel_stage3_mask = unknown_mask & melanoma_mask & df['N'].str.contains('N1|N2|N3', na=False)
            df.loc[mel_stage3_mask, 'STAGE_SIMPLE'] = 'III'
            logger.info(f"Inferred stage III for {mel_stage3_mask.sum()} melanoma patients based on N1-3")
        
        # For melanoma with ulceration or higher thickness, assume at least stage II if no nodes
        if 'T' in df.columns:
            mel_stage2_mask = unknown_mask & melanoma_mask & df['T'].str.contains('T3|T4', na=False)
            if 'N' in df.columns:
                mel_stage2_mask = mel_stage2_mask & (df['N'].str.contains('N0', na=False) | df['N'].isna())
            df.loc[mel_stage2_mask, 'STAGE_SIMPLE'] = 'II'
            logger.info(f"Inferred stage II for {mel_stage2_mask.sum()} melanoma patients based on T3-4, N0")
    
    # Analyze results
    updated_unknown = (df['STAGE_SIMPLE'] == 'Unknown').sum()
    resolved = initial_unknown - updated_unknown
    logger.info(f"Resolved {resolved} Unknown stages ({resolved/initial_unknown:.1%} of initial Unknown)")
    logger.info(f"Final stage distribution: {df['STAGE_SIMPLE'].value_counts().to_dict()}")
    
    # Group early and late stage for analysis
    df['STAGE_GROUP'] = df['STAGE_SIMPLE'].apply(
        lambda s: 'Early' if s in ['0', 'I', 'II'] else
        'Late' if s in ['III', 'IV'] else 'Unknown'
    )
    logger.info(f"Stage groups: {df['STAGE_GROUP'].value_counts().to_dict()}")
    
    return df

def parse_args():
    """
    Parse command line arguments.
    """
    parser = argparse.ArgumentParser(description="Process clinical data for melanoma patients")
    parser.add_argument(
        "--base-path",
        type=str,
        default="/project/orien/data/aws/24PRJ217UVA_IORIG",
        help="Base path to project directory containing clinical data"
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default="output/melanoma_analysis",
        help="Directory for output plots and analysis"
    )
    return parser.parse_args()

def configure_logging(output_dir, log_level=logging.INFO):
    """
    Configure logging for the module.

    Args:
        output_dir (str): Directory to save log file
        log_level (int): Logging level (default: INFO)
    """
    # Create logs directory
    log_dir = os.path.join(output_dir, 'logs')
    os.makedirs(log_dir, exist_ok=True)

    # Set up logging to file
    log_file = os.path.join(log_dir, f'eda_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log')

    # Configure root logger
    logging.basicConfig(
        level=log_level,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()  # Also output to console
        ]
    )

    logger.info(f"Logging configured. Log file: {log_file}")
    return log_file

def filter_melanoma_patients(clinical, output_dir=None):
    """
    Filter clinical data to include only melanoma patients based on the CancerType column.
    This is a critical step for focusing analysis on only melanoma cases.

    Args:
        clinical (pd.DataFrame): Clinical data with CancerType column
        output_dir (str, optional): Directory to save filtered data

    Returns:
        pd.DataFrame: Filtered clinical data containing only melanoma patients
    """
    logger.info("==================== FILTERING FOR MELANOMA PATIENTS ====================")
    logger.info("Applying strict melanoma filter based on histology-derived cancer type")

    # Check if we have CancerType column
    if 'CancerType' not in clinical.columns:
        logger.warning("ERROR: No CancerType column found in clinical data")
        logger.warning("Cannot filter for melanoma patients - check that diagnosis data was properly processed")
        return clinical

    # Print the unique cancer types in the dataset
    cancer_types = clinical['CancerType'].value_counts()
    logger.info(f"Cancer type distribution before filtering: {dict(cancer_types)}")

    # Apply melanoma filter - keep only patients with CancerType == 'Melanoma'
    melanoma_mask = clinical['CancerType'] == 'Melanoma'
    melanoma_patients = clinical[melanoma_mask].copy()

    # Log filtering results
    logger.info(f"Selected {melanoma_mask.sum()} melanoma patients out of {len(clinical)} total patients ({melanoma_mask.sum()/len(clinical):.1%})")

    # Check if we have a reasonable number of melanoma patients
    if melanoma_mask.sum() < 10:
        logger.warning(f"WARNING: Only found {melanoma_mask.sum()} melanoma patients - this seems low")
        logger.warning("Check that histology codes in the diagnosis data are correctly captured")

    # Save filtered data if output directory provided
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(output_dir, 'melanoma_patients.csv')
        melanoma_patients.to_csv(output_file, index=False)
        logger.info(f"Saved melanoma patient dataset to {output_file}")

    # Print some statistics about the melanoma cohort
    if not melanoma_patients.empty:
        if 'Sex' in melanoma_patients.columns:
            sex_dist = melanoma_patients['Sex'].value_counts()
            logger.info(f"Sex distribution in melanoma cohort: {dict(sex_dist)}")

        if 'DiagnosisAge' in melanoma_patients.columns:
            age_stats = melanoma_patients['DiagnosisAge'].describe()
            logger.info(f"Age statistics in melanoma cohort: mean={age_stats['mean']:.1f}, median={age_stats['50%']:.1f}, range=({age_stats['min']:.1f}-{age_stats['max']:.1f})")

        if 'HAS_ICB' in melanoma_patients.columns:
            icb_count = melanoma_patients['HAS_ICB'].sum()
            logger.info(f"ICB treatment in melanoma cohort: {icb_count}/{len(melanoma_patients)} patients ({icb_count/len(melanoma_patients):.1%})")

    return melanoma_patients

def process_clinical_data(dfs, output_dir=None):
    """
    Process clinical data from loaded dataframes to create a unified dataset.
    This function integrates information from multiple sources and standardizes data formats.

    Args:
        dfs (dict): Dictionary of dataframes with clinical data
        output_dir (str, optional): Directory to save processed data

    Returns:
        pd.DataFrame: Processed clinical dataframe with integrated data
    """
    logger.info("==================== PROCESSING CLINICAL DATA ====================")
    logger.info("Integrating patient, diagnosis, vital status, and medication data")

    # Initialize base clinical dataframe using patient data
    if 'patients' not in dfs:
        logger.error("ERROR: No patient data found - cannot create clinical dataframe")
        return pd.DataFrame()

    # Start with patient data (demographics)
    patients = dfs['patients'].copy()

    # Log initial data size and available columns
    logger.info(f"Starting with patient master data: {patients.shape[0]} patients, {patients.shape[1]} columns")
    logger.info(f"Available demographic columns: {', '.join(patients.columns[:10])}...")

    # Step 1: Add diagnosis information if available
    diagnosis_df = None
    if 'diagnosis' in dfs:
        logger.info("STEP 1: Adding diagnosis information to patient data")

        # Print summary of diagnosis data
        diagnosis_df = dfs['diagnosis']
        logger.info(f"Diagnosis data contains {diagnosis_df.shape[0]} records for {diagnosis_df['PATIENT_ID'].nunique()} patients")

        # Add diagnosis information to patient data
        patients = _add_diagnosis_info(patients, diagnosis_df)
        logger.info(f"After adding diagnosis info: {patients.shape[0]} patients, {patients.shape[1]} columns")

        # Check how many patients have cancer type assigned
        if 'CancerType' in patients.columns:
            cancer_type_counts = patients['CancerType'].value_counts()
            logger.info(f"Cancer type distribution: {dict(cancer_type_counts)}")
    else:
        logger.warning("WARNING: No diagnosis data found - cannot determine cancer types")

    # Step 2: Add vital status and survival information if available
    outcome_df = None
    if 'vital' in dfs:
        logger.info("STEP 2: Adding survival information to clinical data")

        # Print summary of vital status data
        vital_df = dfs['vital']
        logger.info(f"Vital status data contains records for {vital_df['PATIENT_ID'].nunique()} patients")

        # Add survival information to patient data
        patients = _add_survival_info(patients, vital_df)
        logger.info(f"After adding survival info: {patients.shape[0]} patients, {patients.shape[1]} columns")

        # Check survival data availability
        if 'OS_TIME' in patients.columns and 'OS_STATUS' in patients.columns:
            os_time_available = patients['OS_TIME'].notna().sum()
            os_events = patients['OS_STATUS'].sum()
            logger.info(f"Survival data available for {os_time_available} patients ({os_time_available/len(patients):.1%})")
            logger.info(f"Number of death events: {os_events} ({os_events/len(patients):.1%})")

        # Use vital status as outcome data
        outcome_df = vital_df
        
        # Check if we need to look for a separate outcome dataframe
        # VitalStatus file might not contain progression/recurrence data
        if 'outcomes' in dfs:
            logger.info("Using dedicated outcomes data")
            outcome_df = dfs['outcomes']
    else:
        logger.warning("WARNING: No vital status data found - cannot calculate survival endpoints")

    # Step 3: Process medications to identify ICB treatments
    if 'medications' in dfs:
        logger.info("STEP 3: Processing medication data to identify ICB treatments")

        # Print summary of medication data
        medications_df = dfs['medications']
        logger.info(f"Medication data contains {medications_df.shape[0]} records for {medications_df['PATIENT_ID'].nunique()} patients")

        # Process medication data to identify ICB treatments
        icb_data = _process_icb_data(medications_df)

        if len(icb_data) > 0:
            # Add ICB data to patients
            patients = patients.merge(icb_data, on='PATIENT_ID', how='left')

            # Create HAS_ICB flag (1 if patient received ICB, 0 otherwise)
            patients['HAS_ICB'] = patients['ICB_STATUS'].notna().astype(int)

            # Print summary of ICB data
            icb_count = patients['HAS_ICB'].sum()
            logger.info(f"Found {icb_count}/{len(patients)} patients ({icb_count/len(patients):.1%}) with ICB treatment")

            # Check ICB class distribution (PD-1, PD-L1, CTLA-4, etc.)
            if 'ICB_CLASS' in patients.columns:
                icb_class_counts = patients[patients['HAS_ICB'] == 1]['ICB_CLASS'].value_counts()
                logger.info(f"ICB class distribution: {dict(icb_class_counts)}")

            # Calculate stage at time of ICB treatment
            logger.info("Calculating stage at time of ICB treatment")
            if 'diagnosis' in dfs:
                diagnosis_df = dfs['diagnosis']
                outcome_df = dfs.get('vital', None)

                # Initialize stage at ICB columns
                patients['StageAtICB'] = 'Unknown'
                patients['StagePrimaryAtICB'] = np.nan
                patients['HistologyAtICB'] = np.nan

                # Process each ICB patient to determine their stage at treatment
                icb_patient_count = patients['HAS_ICB'].sum()
                stage_found_count = 0
                
                # Log available columns in outcome_df for debugging
                if outcome_df is not None:
                    logger.info(f"Available columns in outcome data: {outcome_df.columns.tolist()}")
                    # Check for necessary columns
                    if 'AgeAtProgRecur' not in outcome_df.columns:
                        logger.warning("AgeAtProgRecur column not found in outcome data - will use alternative approach")
                    if 'ProgRecurInd' not in outcome_df.columns:
                        logger.warning("ProgRecurInd column not found in outcome data - recurrence info will be limited")
                
                for idx, row in patients[patients['HAS_ICB'] == 1].iterrows():
                    patient_id = row['PATIENT_ID']
                    icb_start_age = row.get('ICB_START_AGE')
                    
                    # Skip if no ICB start age
                    if pd.isna(icb_start_age):
                        logger.warning(f"No ICB start age available for patient {patient_id}")
                        continue
                    
                    try:
                        # Get patient's stage at ICB - catch errors for individual patients
                        stage, primary_site, histology = get_stage_at_icb(
                            patient_id, row, icb_start_age, diagnosis_df, outcome_df
                        )
                        
                        # Store values
                        patients.loc[idx, 'StageAtICB'] = stage
                        if pd.notna(primary_site):
                            patients.loc[idx, 'StagePrimaryAtICB'] = primary_site
                        if pd.notna(histology):
                            patients.loc[idx, 'HistologyAtICB'] = histology
                        
                        if stage != 'Unknown':
                            stage_found_count += 1
                    except Exception as e:
                        logger.error(f"Error determining stage at ICB for patient {patient_id}: {e}")
                        patients.loc[idx, 'StageAtICB'] = 'Error'

                # Log results
                if icb_patient_count > 0:
                    logger.info(f"Found stage at ICB for {stage_found_count}/{icb_patient_count} patients ({stage_found_count/icb_patient_count:.1%})")
                    stage_dist = patients['StageAtICB'].value_counts().to_dict()
                    logger.info(f"Stage at ICB distribution: {stage_dist}")
            else:
                logger.warning("Cannot calculate stage at ICB: missing diagnosis data")
        else:
            # Create empty ICB columns if no ICB treatments found
            logger.warning("No ICB treatments identified in medication data")
            patients['HAS_ICB'] = 0
            patients['ICB_STATUS'] = 'NO'
            patients['ICB_START_AGE'] = None
            patients['ICB_CLASS'] = None
    else:
        logger.warning("WARNING: No medication data found - cannot identify ICB treatments")

    # Step 4: Create simplified stage column
    logger.info("STEP 4: Creating simplified stage classification")
    patients = create_stage_simple(patients)

    # Save processed data if output directory provided
    if output_dir:
        processed_dir = os.path.join(os.path.dirname(output_dir), 'processed_data')
        os.makedirs(processed_dir, exist_ok=True)
        output_file = os.path.join(processed_dir, 'processed_clinical.csv')
        patients.to_csv(output_file, index=False)
        logger.info(f"Saved processed clinical data to {output_file}")

    # Summary of processed data
    logger.info("==================== CLINICAL DATA PROCESSING SUMMARY ====================")
    logger.info(f"Final clinical dataset: {patients.shape[0]} patients, {patients.shape[1]} columns")
    key_cols = ['CancerType', 'DiagnosisAge', 'OS_TIME', 'OS_STATUS', 'HAS_ICB', 'StageAtICB', 'STAGE_SIMPLE', 'STAGE_GROUP']
    logger.info(f"Key columns added: {', '.join([col for col in key_cols if col in patients.columns])}")

    return patients

def analyze_sequencing_data(melanoma_clinical, sequencing_df, output_dir):
    """
    Analyze tumor sequencing data in relation to melanoma staging and treatment.
    
    Args:
        melanoma_clinical (pd.DataFrame): Clinical data for melanoma patients
        sequencing_df (pd.DataFrame): Tumor sequencing data
        output_dir (str): Output directory for plots and results
    """
    if melanoma_clinical.empty or sequencing_df.empty:
        logger.warning("Empty dataframes provided to analyze_sequencing_data, skipping analysis")
        return
    
    logger.info("==================== ANALYZING SEQUENCING DATA ====================")
    
    # Check for required columns and map to standard names
    required_cols = ['PATIENT_ID']
    cols_to_include = ['PATIENT_ID']
    
    # Map column names to standard names
    column_map = {}
    
    # Check for stage column
    stage_cols = ['STAGE', 'STAGE_SIMPLE', 'StageAtICB']
    stage_col = None
    for col in stage_cols:
        if col in melanoma_clinical.columns:
            stage_col = col
            cols_to_include.append(col)
            column_map[col] = 'STAGE'
            logger.info(f"Using {col} as stage column")
            break
    
    # Check for ICB status column
    icb_cols = ['ICB_STATUS', 'HAS_ICB']
    icb_col = None
    for col in icb_cols:
        if col in melanoma_clinical.columns:
            icb_col = col
            cols_to_include.append(col)
            column_map[col] = 'ICB_STATUS'
            logger.info(f"Using {col} as ICB status column")
            break
    
    # Check for sex/gender column
    sex_cols = ['GENDER', 'Sex', 'SEX']
    sex_col = None
    for col in sex_cols:
        if col in melanoma_clinical.columns:
            sex_col = col
            cols_to_include.append(col)
            column_map[col] = 'GENDER'
            logger.info(f"Using {col} as gender column")
            break
    
    # Check for survival time column
    time_cols = ['OS_TIME', 'OS_MONTHS', 'SurvivalTime']
    time_col = None
    for col in time_cols:
        if col in melanoma_clinical.columns:
            time_col = col
            cols_to_include.append(col)
            column_map[col] = 'OS_TIME'
            logger.info(f"Using {col} as survival time column")
            break
    
    # Check for survival event column
    event_cols = ['OS_EVENT', 'OS_STATUS', 'DeathEvent']
    event_col = None
    for col in event_cols:
        if col in melanoma_clinical.columns:
            event_col = col
            cols_to_include.append(col)
            column_map[col] = 'OS_EVENT'
            logger.info(f"Using {col} as survival event column")
            break
    
    # Select available columns 
    available_cols = [col for col in cols_to_include if col in melanoma_clinical.columns]
    
    if 'PATIENT_ID' not in available_cols:
        logger.error("PATIENT_ID column not found in clinical data, cannot proceed with analysis")
        return
        
    # Create a subset of clinical data with only the columns we need
    clinical_subset = melanoma_clinical[available_cols].copy()
    
    # Rename columns to standard names
    clinical_subset = clinical_subset.rename(columns=column_map)
    
    # Merge sequencing data with clinical data
    seq_clinical = pd.merge(
        clinical_subset,
        sequencing_df,
        on='PATIENT_ID',
        how='inner'
    )
    
    if seq_clinical.empty:
        logger.warning("No matching records found between clinical and sequencing data")
        return
    
    logger.info(f"Found {len(seq_clinical)} sequenced samples for {seq_clinical['PATIENT_ID'].nunique()} unique melanoma patients")
    
    # Count samples per patient
    samples_per_patient = seq_clinical.groupby('PATIENT_ID').size().reset_index(name='sample_count')
    logger.info(f"Average samples per patient: {samples_per_patient['sample_count'].mean():.2f}")
    logger.info(f"Max samples per patient: {samples_per_patient['sample_count'].max()}")
    
    # Count stages with sequencing if stage column exists
    if 'STAGE' in seq_clinical.columns:
        stage_counts = seq_clinical['STAGE'].value_counts()
        logger.info(f"Stage distribution in sequenced samples:")
        for stage, count in stage_counts.items():
            logger.info(f"  - Stage {stage}: {count} samples")
    
    # Analyze sequencing by ICB status if available
    if 'ICB_STATUS' in seq_clinical.columns:
        icb_counts = seq_clinical['ICB_STATUS'].value_counts()
        logger.info(f"ICB status in sequenced samples:")
        for status, count in icb_counts.items():
            logger.info(f"  - {status}: {count} samples")
    
    # If available, analyze key mutations
    mutation_columns = [col for col in seq_clinical.columns if 'mutation' in col.lower()]
    if mutation_columns:
        logger.info(f"Found {len(mutation_columns)} mutation columns in sequencing data")
        for col in mutation_columns:
            mutation_counts = seq_clinical[col].value_counts(dropna=False)
            logger.info(f"Distribution of {col}:")
            for value, count in mutation_counts.items():
                if pd.isna(value):
                    logger.info(f"  - Missing/NA: {count} samples")
                else:
                    logger.info(f"  - {value}: {count} samples")
    else:
        logger.warning("No mutation columns found in sequenced clinical data")
    
    # Create visualizations
    try:
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        # Plot stage distribution if stage available
        if 'STAGE' in seq_clinical.columns and not seq_clinical['STAGE'].isna().all():
            plt.figure(figsize=(10, 6))
            stage_counts.plot(kind='bar')
            plt.title('Stage Distribution in Sequenced Melanoma Samples')
            plt.xlabel('AJCC Stage')
            plt.ylabel('Number of Samples')
            plt.tight_layout()
            plt.savefig(os.path.join(output_dir, 'sequenced_stage_distribution.png'))
            plt.close()
        
        # Plot samples per patient histogram
        plt.figure(figsize=(10, 6))
        sns.histplot(samples_per_patient['sample_count'], bins=range(1, samples_per_patient['sample_count'].max() + 2))
        plt.title('Number of Sequenced Samples per Patient')
        plt.xlabel('Sample Count')
        plt.ylabel('Number of Patients')
        plt.tight_layout()
        plt.savefig(os.path.join(output_dir, 'samples_per_patient.png'))
        plt.close()
        
        logger.info(f"Saved sequencing analysis visualizations to {output_dir}")
    except Exception as e:
        logger.error(f"Error creating sequencing visualizations: {e}")
    
    # Generate a detailed report on melanoma patients with sequencing data
    try:
        # Create a report directory
        report_dir = os.path.join(output_dir, 'reports')
        os.makedirs(report_dir, exist_ok=True)
        
        # Get unique patients and their sample counts
        unique_patients = seq_clinical['PATIENT_ID'].unique()
        logger.info(f"Generating detailed report for {len(unique_patients)} melanoma patients with sequencing data")
        
        # Create summary tables
        patient_info = []
        
        # Get demographics and clinical info for each patient
        for patient_id in unique_patients:
            patient_data = seq_clinical[seq_clinical['PATIENT_ID'] == patient_id].iloc[0].copy()
            samples_count = samples_per_patient[samples_per_patient['PATIENT_ID'] == patient_id]['sample_count'].values[0]
            
            # Get the key information for the patient
            patient_summary = {
                'PATIENT_ID': patient_id,
                'SequencedSampleCount': samples_count
            }
            
            # Add demographic info if available
            for col_name, standard_name in [
                ('GENDER', 'Gender'), 
                ('STAGE', 'Stage'),
                ('ICB_STATUS', 'ICB_Status')
            ]:
                if col_name in patient_data:
                    patient_summary[standard_name] = patient_data[col_name]
            
            # Add survival info if available
            if 'OS_TIME' in patient_data and 'OS_EVENT' in patient_data:
                patient_summary['OS_Months'] = patient_data['OS_TIME']
                patient_summary['DeathEvent'] = patient_data['OS_EVENT']
            
            # Add mutation info if available
            for col in mutation_columns:
                if col in patient_data:
                    mutation_name = col.replace('_mutation', '').upper()
                    value = patient_data[col]
                    patient_summary[f"{mutation_name}_Mutation"] = "Unknown" if pd.isna(value) else str(value)
            
            patient_info.append(patient_summary)
        
        # Create a dataframe with all patient info
        patient_report = pd.DataFrame(patient_info)
        
        # Generate summary statistics
        report_summary = {
            "Total Melanoma Patients": len(melanoma_clinical),
            "Melanoma Patients with Sequencing": len(unique_patients),
            "Percentage with Sequencing": f"{len(unique_patients)/len(melanoma_clinical):.1%}",
            "Total Sequenced Samples": len(seq_clinical),
            "Average Samples Per Patient": f"{samples_per_patient['sample_count'].mean():.2f}"
        }
        
        # Add gender distribution if available
        if 'Gender' in patient_report.columns:
            gender_counts = patient_report['Gender'].value_counts()
            for gender, count in gender_counts.items():
                report_summary[f"Gender_{gender}"] = f"{count} ({count/len(patient_report):.1%})"
        
        # Add stage distribution if available
        if 'Stage' in patient_report.columns:
            stage_groups = {
                'Early': ['Stage 0', 'Stage I', 'Stage IA', 'Stage IB', 'Stage IS', 'Stage II', 'Stage IIA', 'Stage IIB', 'Stage IIC'],
                'Late': ['Stage III', 'Stage IIIA', 'Stage IIIB', 'Stage IIIC', 'Stage IIID', 'Stage IV', 'Stage IVA', 'Stage IVB', 'Stage IVC']
            }
            
            for group, stages in stage_groups.items():
                count = patient_report[patient_report['Stage'].isin(stages)].shape[0]
                report_summary[f"StageGroup_{group}"] = f"{count} ({count/len(patient_report):.1%})"
            
            # Count unknown stages
            unknown_count = patient_report[~patient_report['Stage'].isin(sum(stage_groups.values(), []))].shape[0]
            report_summary[f"StageGroup_Unknown"] = f"{unknown_count} ({unknown_count/len(patient_report):.1%})"
        
        # Add ICB treatment info if available
        if 'ICB_Status' in patient_report.columns:
            icb_counts = patient_report['ICB_Status'].value_counts()
            for status, count in icb_counts.items():
                if pd.isna(status):
                    status = "Unknown"
                report_summary[f"ICB_{status}"] = f"{count} ({count/len(patient_report):.1%})"
        
        # Add mutation statistics if available
        for col in [c for c in patient_report.columns if 'Mutation' in c]:
            mutation_counts = patient_report[col].value_counts(dropna=False)
            for value, count in mutation_counts.items():
                if pd.isna(value):
                    value = "Unknown"
                report_summary[f"{col}_{value}"] = f"{count} ({count/len(patient_report):.1%})"
        
        # Create summary dataframe
        summary_df = pd.DataFrame(list(report_summary.items()), columns=['Metric', 'Value'])
        
        # Save the reports
        patient_report.to_csv(os.path.join(report_dir, 'melanoma_sequenced_patients_report.csv'), index=False)
        summary_df.to_csv(os.path.join(report_dir, 'melanoma_sequencing_summary.csv'), index=False)
        
        # Create Excel report with both tables
        try:
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            
            # Summary sheet
            summary_sheet = wb.active
            summary_sheet.title = "Summary"
            
            # Add title
            summary_sheet['A1'] = "Melanoma Sequencing Analysis Summary"
            summary_sheet['A1'].font = Font(bold=True, size=14)
            summary_sheet.merge_cells('A1:B1')
            
            # Add summary data
            for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False), 3):
                for c_idx, value in enumerate(row, 1):
                    summary_sheet.cell(row=r_idx, column=c_idx, value=value)
            
            # Patient details sheet
            patient_sheet = wb.create_sheet("Patient Details")
            
            # Add title
            patient_sheet['A1'] = "Melanoma Patients with Sequencing Data"
            patient_sheet['A1'].font = Font(bold=True, size=14)
            patient_sheet.merge_cells('A1:F1')
            
            # Add patient data
            for r_idx, row in enumerate(dataframe_to_rows(patient_report, index=False), 3):
                for c_idx, value in enumerate(row, 1):
                    patient_sheet.cell(row=r_idx, column=c_idx, value=value)
            
            # Save the workbook
            excel_file = os.path.join(report_dir, 'melanoma_sequencing_report.xlsx')
            wb.save(excel_file)
            logger.info(f"Saved comprehensive Excel report to {excel_file}")
        except Exception as e:
            logger.warning(f"Could not create Excel report: {e}")
        
        logger.info(f"Saved detailed reports to {report_dir}")
        
    except Exception as e:
        logger.error(f"Error generating detailed sequencing report: {e}")
    
    # Save the merged data for further analysis
    seq_clinical_file = os.path.join(output_dir, 'sequenced_melanoma_clinical.csv')
    seq_clinical.to_csv(seq_clinical_file, index=False)
    logger.info(f"Saved sequenced melanoma clinical data to {seq_clinical_file}")
    
    # Analyze mutations vs outcomes if both survival columns exist
    if 'OS_TIME' in seq_clinical.columns and 'OS_EVENT' in seq_clinical.columns:
        mutation_summary = analyze_mutations_vs_outcomes(seq_clinical, output_dir)
    
    return seq_clinical

def analyze_mutations_vs_outcomes(seq_clinical, output_dir):
    """
    Analyze the relationship between mutation status and clinical outcomes
    for melanoma patients with sequencing data.
    
    Args:
        seq_clinical (pd.DataFrame): DataFrame containing merged clinical and sequencing data
        output_dir (str): Output directory for saving plots and results
    """
    if seq_clinical.empty:
        logger.warning("Empty sequenced clinical data provided, skipping mutation outcome analysis")
        return
    
    logger.info("==================== ANALYZING MUTATIONS VS OUTCOMES ====================")
    
    # Identify mutation columns
    mutation_columns = [col for col in seq_clinical.columns if 'mutation' in col.lower()]
    if not mutation_columns:
        logger.warning("No mutation columns found in sequenced clinical data")
        return
    
    logger.info(f"Analyzing {len(mutation_columns)} mutation types vs clinical outcomes")
    
    # Check for required columns
    required_cols = ['OS_TIME', 'OS_EVENT']
    if not all(col in seq_clinical.columns for col in required_cols):
        logger.warning(f"Required columns {required_cols} not found in dataset, skipping survival analysis")
        return
    
    # Generate reports directory
    mutation_dir = os.path.join(output_dir, 'mutation_analysis')
    os.makedirs(mutation_dir, exist_ok=True)
    
    # Analyze each mutation type
    mutation_summary = {}
    
    for mutation_col in mutation_columns:
        mutation_name = mutation_col.replace('_mutation', '').upper()
        logger.info(f"Analyzing {mutation_name} mutation")
        
        # Clean the mutation data - handle different formats
        if seq_clinical[mutation_col].dtype == 'object':
            # Convert text values to boolean if needed
            seq_clinical[mutation_col] = seq_clinical[mutation_col].map(
                lambda x: pd.NA if pd.isna(x) else 
                         (True if str(x).lower() in ['yes', 'true', '1', 'positive', 'mutated'] else 
                          (False if str(x).lower() in ['no', 'false', '0', 'negative', 'wild-type'] else x))
            )
        
        # Count mutation prevalence
        mutation_counts = seq_clinical[mutation_col].value_counts(dropna=False)
        mutated_count = mutation_counts.get(True, 0)
        wildtype_count = mutation_counts.get(False, 0)
        unknown_count = seq_clinical[mutation_col].isna().sum()
        
        logger.info(f"{mutation_name} mutation status:")
        logger.info(f"  - Mutated: {mutated_count} patients")
        logger.info(f"  - Wild-type: {wildtype_count} patients")
        logger.info(f"  - Unknown: {unknown_count} patients")
        
        # Skip if not enough data
        if mutated_count < 5 or wildtype_count < 5:
            logger.warning(f"Not enough patients in each group for {mutation_name} analysis, skipping")
            continue
        
        # Analyze mutation vs stage
        if 'STAGE' in seq_clinical.columns:
            try:
                # Create cross-tabulation
                stage_vs_mutation = pd.crosstab(
                    seq_clinical['STAGE'], 
                    seq_clinical[mutation_col],
                    margins=True
                )
                
                # Save to file
                stage_file = os.path.join(mutation_dir, f'{mutation_name}_stage_distribution.csv')
                stage_vs_mutation.to_csv(stage_file)
                logger.info(f"Saved {mutation_name} stage distribution to {stage_file}")
                
                # Plot
                plt.figure(figsize=(10, 6))
                sns.countplot(data=seq_clinical.dropna(subset=[mutation_col, 'STAGE']), 
                             x='STAGE', hue=mutation_col)
                plt.title(f'{mutation_name} Mutation by Stage')
                plt.xlabel('AJCC Stage')
                plt.ylabel('Patient Count')
                plt.tight_layout()
                plt.savefig(os.path.join(mutation_dir, f'{mutation_name}_stage_distribution.png'))
                plt.close()
            except Exception as e:
                logger.error(f"Error analyzing {mutation_name} vs stage: {e}")
        
        # Survival analysis for this mutation
        try:
            # Filter data for survival analysis
            survival_data = seq_clinical.dropna(subset=[mutation_col, 'OS_TIME', 'OS_EVENT'])
            
            # Create survival plots only if we have sufficient data
            if len(survival_data) >= 10:
                # Plot Kaplan-Meier curves stratified by mutation status
                plt.figure(figsize=(10, 6))
                
                # Create KM model for mutated and wild-type
                kmf_mutated = KaplanMeierFitter()
                kmf_wildtype = KaplanMeierFitter()
                
                # Get indices for each group
                mutated_mask = survival_data[mutation_col] == True
                wildtype_mask = survival_data[mutation_col] == False
                
                # Fit and plot if we have patients in each group
                if sum(mutated_mask) >= 5 and sum(wildtype_mask) >= 5:
                    T_mutated = survival_data.loc[mutated_mask, 'OS_TIME']
                    E_mutated = survival_data.loc[mutated_mask, 'OS_EVENT']
                    
                    T_wildtype = survival_data.loc[wildtype_mask, 'OS_TIME']
                    E_wildtype = survival_data.loc[wildtype_mask, 'OS_EVENT']
                    
                    # Fit KM curves
                    kmf_mutated.fit(T_mutated, E_mutated, label=f'{mutation_name} Mutated (n={sum(mutated_mask)})')
                    kmf_wildtype.fit(T_wildtype, E_wildtype, label=f'{mutation_name} Wild-type (n={sum(wildtype_mask)})')
                    
                    # Plot
                    kmf_mutated.plot(show_censors=True, ci_show=True)
                    kmf_wildtype.plot(show_censors=True, ci_show=True)
                    
                    # Calculate log-rank p-value
                    results = logrank_test(T_mutated, T_wildtype, E_mutated, E_wildtype)
                    p_value = results.p_value
                    
                    # Format the p-value with appropriate precision
                    if p_value < 0.001:
                        p_value_str = "p < 0.001"
                    elif p_value < 0.01:
                        p_value_str = f"p = {p_value:.3f}"
                    else:
                        p_value_str = f"p = {p_value:.2f}"
                    
                    # Add p-value to the plot
                    plt.text(0.7, 0.2, f"Log-rank test:\n{p_value_str}", 
                             transform=plt.gca().transAxes, bbox=dict(facecolor='white', alpha=0.8))
                    
                    # Add HR calculation using Cox model if available
                    try:
                        from lifelines import CoxPHFitter
                        
                        # Create a dataset for Cox analysis
                        cox_data = survival_data[['OS_TIME', 'OS_EVENT', mutation_col]].copy()
                        cox_data['mutation'] = cox_data[mutation_col].astype(int)  # Convert to 0/1
                        
                        # Fit Cox model
                        cph = CoxPHFitter()
                        cph.fit(cox_data, duration_col='OS_TIME', event_col='OS_EVENT', formula="mutation")
                        
                        # Extract HR and CI
                        hr = np.exp(cph.params_[0])
                        hr_lower = np.exp(cph.confidence_intervals_.loc['mutation', 'coef lower 95%'])
                        hr_upper = np.exp(cph.confidence_intervals_.loc['mutation', 'coef upper 95%'])
                        
                        # Add HR to the plot
                        plt.text(0.7, 0.1, f"HR = {hr:.2f} (95% CI: {hr_lower:.2f}-{hr_upper:.2f})", 
                                transform=plt.gca().transAxes, bbox=dict(facecolor='white', alpha=0.8))
                        
                        # Save the Cox model results
                        cox_file = os.path.join(mutation_dir, f'{mutation_name}_cox_results.csv')
                        cph.summary.to_csv(cox_file)
                        
                        # Store in summary
                        mutation_summary[mutation_name] = {
                            'p_value': p_value,
                            'hazard_ratio': hr,
                            'hr_ci_lower': hr_lower,
                            'hr_ci_upper': hr_upper,
                            'mutated_n': sum(mutated_mask),
                            'wildtype_n': sum(wildtype_mask)
                        }
                    except Exception as e:
                        logger.error(f"Error calculating HR for {mutation_name}: {e}")
                        mutation_summary[mutation_name] = {
                            'p_value': p_value,
                            'mutated_n': sum(mutated_mask),
                            'wildtype_n': sum(wildtype_mask)
                        }
                    
                    plt.title(f'Overall Survival by {mutation_name} Mutation Status')
                    plt.xlabel('Time (months)')
                    plt.ylabel('Survival Probability')
                    plt.ylim(0, 1.05)
                    plt.tight_layout()
                    plt.savefig(os.path.join(mutation_dir, f'{mutation_name}_survival.png'))
                    plt.close()
                    
                    logger.info(f"Completed survival analysis for {mutation_name} mutation")
                else:
                    logger.warning(f"Not enough patients in each group for {mutation_name} survival analysis")
            else:
                logger.warning(f"Not enough survival data for {mutation_name} analysis")
        except Exception as e:
            logger.error(f"Error in survival analysis for {mutation_name}: {e}")
    
    # Create a summary table
    if mutation_summary:
        summary_df = pd.DataFrame.from_dict(mutation_summary, orient='index')
        summary_file = os.path.join(mutation_dir, 'mutation_survival_summary.csv')
        summary_df.to_csv(summary_file)
        logger.info(f"Saved mutation survival summary to {summary_file}")
        
        # Create a forest plot if we have hazard ratios
        if 'hazard_ratio' in summary_df.columns and len(summary_df) > 1:
            try:
                # Sort by hazard ratio
                summary_df = summary_df.sort_values('hazard_ratio')
                
                plt.figure(figsize=(12, len(summary_df) * 0.8 + 2))
                
                # Plot the hazard ratios and CIs
                y_pos = np.arange(len(summary_df))
                plt.errorbar(
                    x=summary_df['hazard_ratio'],
                    y=y_pos,
                    xerr=[
                        summary_df['hazard_ratio'] - summary_df['hr_ci_lower'],
                        summary_df['hr_ci_upper'] - summary_df['hazard_ratio']
                    ],
                    fmt='o',
                    capsize=5
                )
                
                # Add vertical line at HR=1
                plt.axvline(x=1, color='gray', linestyle='--')
                
                # Add mutation names
                plt.yticks(y_pos, summary_df.index)
                
                # Add labels and title
                plt.xlabel('Hazard Ratio (log scale)')
                plt.title('Forest Plot of Mutation Hazard Ratios')
                
                # Use log scale for x-axis
                plt.xscale('log')
                
                # Add p-values as annotations
                for i, (_, row) in enumerate(summary_df.iterrows()):
                    if row['p_value'] < 0.001:
                        p_str = "p < 0.001"
                    elif row['p_value'] < 0.01:
                        p_str = f"p = {row['p_value']:.3f}"
                    else:
                        p_str = f"p = {row['p_value']:.2f}"
                    
                    plt.text(
                        max(summary_df['hr_ci_upper']) * 1.2,
                        i,
                        p_str,
                        va='center'
                    )
                
                plt.tight_layout()
                plt.savefig(os.path.join(mutation_dir, 'mutation_forest_plot.png'))
                plt.close()
                
                logger.info("Created forest plot for mutation hazard ratios")
            except Exception as e:
                logger.error(f"Error creating forest plot: {e}")

    return mutation_summary

def analyze_survival(df, output_dir):
    """
    Analyze survival data and generate survival statistics.
    
    Args:
        df (pd.DataFrame): Clinical data with survival information
        output_dir (str): Directory to save results
    """
    logger.info("==================== ANALYZING SURVIVAL DATA ====================")

    try:
        # Create survival analysis directory
        survival_dir = os.path.join(output_dir, 'survival')
        os.makedirs(survival_dir, exist_ok=True)

        # Check for required columns
        if 'OS_TIME' not in df.columns or 'OS_EVENT' not in df.columns:
            logger.warning("Required survival columns (OS_TIME, OS_EVENT) not found")
            return

        # Prepare data
        survival_data = df[['OS_TIME', 'OS_EVENT']].copy()
        survival_data['OS_TIME'] = pd.to_numeric(survival_data['OS_TIME'], errors='coerce')
        survival_data['OS_EVENT'] = pd.to_numeric(survival_data['OS_EVENT'], errors='coerce')
        
        # Drop rows with missing survival data
        survival_data = survival_data.dropna()
        
        if len(survival_data) == 0:
            logger.warning("No valid survival data available after cleaning")
            return

        # Basic survival statistics
        stats = {
            'Total Patients': len(survival_data),
            'Events': int(survival_data['OS_EVENT'].sum()),
            'Median Survival Time': survival_data['OS_TIME'].median(),
            'Mean Survival Time': survival_data['OS_TIME'].mean(),
            'Min Survival Time': survival_data['OS_TIME'].min(),
            'Max Survival Time': survival_data['OS_TIME'].max()
        }
        
        # Save survival statistics
        stats_df = pd.DataFrame.from_dict(stats, orient='index', columns=['Value'])
        stats_file = os.path.join(survival_dir, 'survival_statistics.csv')
        stats_df.to_csv(stats_file)
        logger.info(f"Saved survival statistics to {stats_file}")

        # Additional analyses by subgroups if available
        if 'STAGE' in df.columns:
            try:
                stage_groups = df.groupby('STAGE')
                stage_stats = {}
                
                for stage, group in stage_groups:
                    group_data = group[['OS_TIME', 'OS_EVENT']].dropna()
                    if len(group_data) > 0:
                        stage_stats[stage] = {
                            'N': len(group_data),
                            'Events': int(group_data['OS_EVENT'].sum()),
                            'Median Survival': group_data['OS_TIME'].median()
                        }
                
                stage_stats_df = pd.DataFrame.from_dict(stage_stats, orient='index')
                stage_file = os.path.join(survival_dir, 'survival_by_stage.csv')
                stage_stats_df.to_csv(stage_file)
                logger.info(f"Saved survival statistics by stage to {stage_file}")
            except Exception as e:
                logger.error(f"Error analyzing survival by stage: {e}")

        if 'ICB_STATUS' in df.columns or 'HAS_ICB' in df.columns:
            try:
                icb_col = 'ICB_STATUS' if 'ICB_STATUS' in df.columns else 'HAS_ICB'
                icb_groups = df.groupby(icb_col)
                icb_stats = {}
                
                for icb_status, group in icb_groups:
                    group_data = group[['OS_TIME', 'OS_EVENT']].dropna()
                    if len(group_data) > 0:
                        icb_stats[icb_status] = {
                            'N': len(group_data),
                            'Events': int(group_data['OS_EVENT'].sum()),
                            'Median Survival': group_data['OS_TIME'].median()
                        }
                
                icb_stats_df = pd.DataFrame.from_dict(icb_stats, orient='index')
                icb_file = os.path.join(survival_dir, 'survival_by_icb.csv')
                icb_stats_df.to_csv(icb_file)
                logger.info(f"Saved survival statistics by ICB status to {icb_file}")
            except Exception as e:
                logger.error(f"Error analyzing survival by ICB status: {e}")

    except Exception as e:
        logger.error(f"Error in survival analysis: {e}")
        logger.error(traceback.format_exc())

def main(args):
    """
    Main execution function for the clinical data processing pipeline.
    
    Args:
        args: Parsed command-line arguments
    """
    try:
        # Configure logging
        configure_logging(args.output_dir)

        # Set up output directories
        output_dir = setup_output_dirs(args.base_path)

        # Load clinical data
        dfs = load_clinical_data(args.base_path)
        if not dfs:
            logger.error("No clinical data loaded. Exiting.")
            return
            
        # Process clinical data
        clinical = process_clinical_data(dfs, output_dir)

        # Filter for melanoma patients
        melanoma_clinical = filter_melanoma_patients(clinical, output_dir)
        
        # Filter for melanoma patients with sequencing data
        logger.info("==================== FILTERING FOR MELANOMA PATIENTS WITH SEQUENCING DATA ====================")
        
        sequenced_melanoma = melanoma_clinical.copy()
        if 'tumor_sequencing' in dfs and not dfs['tumor_sequencing'].empty:
            sequencing_df = dfs['tumor_sequencing']
            
            # Get patient IDs with sequencing data
            if 'PATIENT_ID' in sequencing_df.columns:
                sequenced_patients = sequencing_df['PATIENT_ID'].unique()
                logger.info(f"Found {len(sequenced_patients)} patients with sequencing data")
                
                # Filter melanoma patients to only those with sequencing data
                sequenced_melanoma = melanoma_clinical[melanoma_clinical['PATIENT_ID'].isin(sequenced_patients)].copy()
                logger.info(f"Filtered to {len(sequenced_melanoma)} melanoma patients with sequencing data " +
                           f"(out of {len(melanoma_clinical)} total melanoma patients)")
                
                # Use this as our primary dataset from now on
                melanoma_clinical = sequenced_melanoma
                
                # Save the filtered dataset
                sequenced_melanoma_file = os.path.join(output_dir, 'sequenced_melanoma_patients.csv')
                sequenced_melanoma.to_csv(sequenced_melanoma_file, index=False)
                logger.info(f"Saved sequenced melanoma patient dataset to {sequenced_melanoma_file}")
                
                # Analyze sequencing data
                seq_clinical = analyze_sequencing_data(sequenced_melanoma, sequencing_df, output_dir)
            else:
                logger.warning("No PATIENT_ID column found in sequencing data - cannot filter melanoma patients")
        else:
            logger.warning("No sequencing data found - cannot filter melanoma patients by sequencing status")

        # Generate demographic plots
        if not melanoma_clinical.empty:
            plot_demographics(melanoma_clinical, output_dir)

        # Analyze survival data
        if 'OS_TIME' in melanoma_clinical.columns and 'OS_EVENT' in melanoma_clinical.columns:
            analyze_survival(melanoma_clinical, output_dir)
            plot_survival_curves(melanoma_clinical, output_dir)

        # Save final processed data
        save_processed_data(melanoma_clinical, os.path.join(args.base_path, 'processed_data'))

        logger.info("Clinical data processing completed successfully.")
        
    except Exception as e:
        logger.error(f"Error in main execution: {e}")
        logger.error(traceback.format_exc())
        raise

# Check if this is being run as a script vs imported as a module
if __name__ == "__main__":
    try:
        main(parse_args())
    except Exception as e:
        logger.error(f"Error in main execution: {e}")
        logger.error(traceback.format_exc())
        raise